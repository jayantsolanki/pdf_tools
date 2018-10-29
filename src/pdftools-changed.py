#main.py
from PyPDF2 import PdfFileWriter, PdfFileReader
from PyPDF2.pdf import ContentStream
from PyPDF2.generic import NameObject, NumberObject, ArrayObject, TextStringObject, RectangleObject, DictionaryObject, BooleanObject, IndirectObject
from PyPDF2.utils import isString, b_
from openpyxl import load_workbook # library function for reading the excel file
import re
import sys # for command line arguments
from argparse import ArgumentParser
import traceback
# this library has been added for only to fetch the bookmarks
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import os
import uuid 
###################End of Libraries#######################

sys.setrecursionlimit(10000) # 10000 is an example, try with different values if current number doesn't suit

# setting up the command-line parameter structure for the python script
parser = ArgumentParser(description='Python based toolkit for repairing define.pdf')
parser.add_argument("-c", "--configFile", help="Takes the path for the config spreadsheet", required=False)#optional
parser.add_argument("-s", "--specFile", help="Takes the specification excel file", default = "", required=False)#optional
parser.add_argument("-i", "--inFile", help="Takes input pdf file for modification", required=True)#required
parser.add_argument("-o", "--outFile", help="Writes the pdf file with the given name", default = "", required=False)#optional
parser.add_argument("-l", "--logFile", help="Writes log messages on a file", default = "", required=False)#optional
parser.add_argument("-v", "--verbose", help="Enable verbose python log", default = False, required=False)#optional
args = parser.parse_args()
extLinks = str(uuid.uuid4())+".tsv" #name for the spreadsheet which will contain the links identified in the pdf document
#### just for debugging purposes###############
# if args.debug:
#     print ("NOTE: *** "+ str(args)+ " ***")
############################################ Important methods ##############################################
#######################################################################################
def progExit():
    sys.stdout.flush()
    try:
        os.remove(extLinks)
    except:
        pass
    sys.exit()

def checkForAnaspec(outlines):
    for item in outlines:
        if(isinstance(item, list)):
            continue
        if (re.search(r'Analysis Results Metadata', str(item['/Title']), re.M|re.I)):
            return True
    return False

### Reads bookmarks and return a tuple array ###
def get_toc(pdf_path):
    infile = open(pdf_path, 'rb')
    parser = PDFParser(infile)
    document = PDFDocument(parser)

    toc = list()
    for (level,title,dest,a,structelem) in document.get_outlines():
        toc.append((level, title))

    return toc

def readSpecExcel(filename = None):
    filemap = {}# dictionary for storing the filemaping
    subLinks = {}# dictionary for storing the sublinks for the Table and Figures
    try:
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb.active

        rows = ws.rows
        for row in rows:

            if (row[0].value!=None):
                if (str(row[0].value).strip() not in subLinks):
                    subLinks[str(row[0].value).strip()] = []
                filemap[str(row[0].value).strip()] = str(row[3].value).strip()
                subLinks[str(row[0].value).strip()].append(str(row[6].value).strip())
    except Exception as e:
        print ("ERROR: *** ANASPEC file is missing in the working directory provided by the User ***")
        progExit()
    return filemap, subLinks

def readConfExcel(filename = None):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active
    oprMap = {}# dictionary for storing the operations
    rows = ws.rows
    for row in rows:

        if (row[1].value!=None and row[2].value!='Category'):
            if (str(row[2].value).strip() not in oprMap):
                oprMap[str(row[2].value).strip()] = {}
            if(str(row[4].value).strip() == 'Yes'):
                oprMap[str(row[2].value).strip()][str(row[1].value).strip()] = True
            else:
                oprMap[str(row[2].value).strip()][str(row[1].value).strip()] = False
    return oprMap
    # for row in rows:

    #   if (row[0].value!=None and row[0].value!='OpId'):
    #       # print str(row[2].value).strip()
    #       if(str(row[2].value).strip() == 'Yes'):
    #           oprMap[str(row[0].value).strip()] = True
    #       else:
    #           oprMap[str(row[0].value).strip()] = False
    # return oprMap

def parseContent(page, debug=False):
    content = page['/Contents'].getObject()
    content = ContentStream(content, page)#creating contentstream class instance
    linkNames = []
    flag = False
    for operands,operator in content.operations:
        try:
            if operator == b_('rg') and operands == [0,0,1]:#links only
                flag = True
                text = operands[0]
            if operator == b_('Tj') and flag: # first parameter, word to be searched, second paramter, word series were word may be present
                # text = operands[0]
                flag = False
                linkNames.append(operands[0])

        except:
            print ('WARNING: *** problem in deciphering the stream page ***')
    # splitContent =  content.splitlines();
    # for text in splitlines:
    # linkNames = [splitContent[index+1]for index, search in enumerate(splitContent) if search == '0 0 1 rg']# looking for text which are blue in color, that is hyperlinked
    # print (splitContent[indices[:]])
    # print (indices)
    #add some error control here, what if nothing is found
    return linkNames
#######################################################################################
def removeExtLinks(page, objectId=None, debug=False):
    """
    Removes External links from this output.
    """
    # pages = self.getObject(self._pages)['/Kids']
    # for page in pages:
    #     pageRef = self.getObject(page)
    # page = input.getPage(pagenum)
    if "/Annots" in page:
        temp = page["/Annots"][:]
        if (objectId == None):#delete all the external links
            del page['/Annots']
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                obj = annot.getObject()
                if(obj['/A']['/S'] == '/GoToR' or obj['/A']['/S'] == '/Launch' or obj['/A']['/S'] == '/URI'):#look for external links only
                    if debug:
                        print ("NOTE: *** Deleted object id "+str(annot.idnum)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
            if debug:
                print ("NOTE: *** Deleted all the external links ***")
        else:#delete the particular indirect object, or the external link
            # print page
            del page["/Annots"]
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                if (annot.idnum == objectId):
                    if debug:
                        print ("NOTE: *** Deleted object id "+str(objectId)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
    return page

def removeIntLinks(page, objectId=None, debug=False): # remove internal links
    """
    Removes Internal links from this output.
    """
    if "/Annots" in page:
        temp = page["/Annots"][:]
        if (objectId == None):#delete all the external links
            del page['/Annots']
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                obj = annot.getObject()
                if(obj['/A']['/S'] == '/GoTo'):#look for internal links only
                    if debug:
                        print ("NOTE: *** Deleted object id "+str(annot.idnum)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
            if debug:
                print ("NOTE: *** Deleted all the external links ***")
        else:#delete the particular indirect object, or the external link
            # print page
            del page["/Annots"]
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                if (annot.idnum == objectId):
                    if debug:
                        print ("NOTE: *** Deleted object id "+ str(objectId)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
    return page

def addLAUNCH(output, page, path, rect, border=None, debug=False):
    """
    Add an Launch type of external Link from a rectangular area to the specified pdf.
    This method is different from what I provided in the pdf.py library
    Here you can directly send the page you want to edit instead of the page number
    Helps in doing many things at a time without going through the need of access the file sequentially
    This uses the basic structure of AddUri

    :param int pagenum: index of the page on which to place the URI action.
    :param int uri: string -- uri of resource to link to.
    :param rect: :class:`RectangleObject<PyPDF2.generic.RectangleObject>` or array of four
        integers specifying the clickable rectangular area
        ``[xLL, yLL, xUR, yUR]``, or string in the form ``"[ xLL yLL xUR yUR ]"``.
    :param border: if provided, an array describing border-drawing
        properties. See the PDF spec for details. No border will be
        drawn if this argument is omitted or given [0,0,0].
    -Jayant Solanki
    """
    C = [0,0,1]#border color
    CObject = ArrayObject([NameObject(n) for n in C])
    if border is not None:
        borderArr = [NameObject(n) for n in border[:3]]
        if len(border) == 4:
            dashPattern = ArrayObject([NameObject(n) for n in border[3]])
            borderArr.append(dashPattern)
    else:
        borderArr = [NumberObject(0)] * 3

    if isString(rect):
        rect = NameObject(rect)
    elif isinstance(rect, RectangleObject):
        pass
    else:
        rect = RectangleObject(rect)
    lnk2 = DictionaryObject()
    lnk2.update({
    NameObject('/S'): NameObject('/Launch'),
    NameObject('/F'): TextStringObject(path),
    NameObject('/NewWindow') : BooleanObject('true'),
    # NameObject('/D') : ArrayObject([NameObject('0/XYZ 0 0 0')])#inherent zoom, last number should be zero
    # NameObject('/D') : ArrayObject([NumberObject(0), NameObject('/XYZ'), NameObject(0), NameObject(0), NumberObject(0)])#inherent zoom, last number should be zero
    });
    lnk = DictionaryObject()
    lnk.update({
    NameObject('/Type'): NameObject('/Annot'),
    NameObject('/Subtype'): NameObject('/Link'),
    # NameObject('/P'): pageLink,
    NameObject('/Rect'): rect,
    NameObject('/H'): NameObject('/I'),
    NameObject('/Border'): ArrayObject(borderArr),
    NameObject('/C'): CObject,
    # NameObject('/NewWindow') : BooleanObject('true'),
    NameObject('/A'): lnk2
    })
    lnkRef = output._addObject(lnk)# creating new object for indirect reference
    if "/Annots" in page:
        page['/Annots'].append(lnkRef)
    else:
        page[NameObject('/Annots')] = ArrayObject([lnkRef])
    return output, page

def addGOTOR(output, page, path, rect, border=None, destination = 0, debug=False):
    """
    Add an GoToR type of external Link from a rectangular area to the specified pdf.
    This method is different from what I provided in the pdf.py library
    Here you can directly send the page you want to edit instead of the page number
    Helps in doing many things at a time without going through the need of access the file sequentially
    This uses the basic structure of AddUri

    :param int pagenum: index of the page on which to place the URI action.
    :param int uri: string -- uri of resource to link to.
    :param rect: :class:`RectangleObject<PyPDF2.generic.RectangleObject>` or array of four
        integers specifying the clickable rectangular area
        ``[xLL, yLL, xUR, yUR]``, or string in the form ``"[ xLL yLL xUR yUR ]"``.
    :param border: if provided, an array describing border-drawing
        properties. See the PDF spec for details. No border will be
        drawn if this argument is omitted or given [0,0,0].
    -Jayant Solanki
    """
    C = [0,0,1]#border color
    CObject = ArrayObject([NameObject(n) for n in C])
    if border is not None:
        borderArr = [NameObject(n) for n in border[:3]]
        if len(border) == 4:
            dashPattern = ArrayObject([NameObject(n) for n in border[3]])
            borderArr.append(dashPattern)
    else:
        borderArr = [NumberObject(0)] * 3

    if isString(rect):
        rect = NameObject(rect)
    elif isinstance(rect, RectangleObject):
        pass
    else:
        rect = RectangleObject(rect)
    lnk2 = DictionaryObject()
    lnk2.update({
    # NameObject('/NewWindow') : BooleanObject('true'),
    NameObject('/S'): NameObject('/GoToR'),
    # NameObject('/NewWindow') : BooleanObject('true'),
    NameObject('/F'): TextStringObject(path), # had to do like this other wise not working
    NameObject('/NewWindow true /D') : ArrayObject([NumberObject(int(destination)-1), NameObject('/XYZ'), NameObject(0), NameObject(10000), NameObject(0.0)])#inherent zoom, last number should be zero
    # NameObject('/D'): TextStringObject('12')
    });
    lnk = DictionaryObject()
    lnk.update({
    NameObject('/Type'): NameObject('/Filespec'),
    NameObject('/Subtype'): NameObject('/Link'),
    # NameObject('/P'): pageLink,
    NameObject('/Rect'): rect,
    NameObject('/H'): NameObject('/I'),
    NameObject('/Border'): ArrayObject(borderArr),
    NameObject('/C'): CObject,
    NameObject('/A'): lnk2
    })
    lnkRef = output._addObject(lnk)# creating new object for indirect reference
    if "/Annots" in page:
        page['/Annots'].append(lnkRef)
    else:
        page[NameObject('/Annots')] = ArrayObject([lnkRef])
    return output, page

def addGOTO(output, page, pageId, rect, border=None, debug=False):
    """
    Add an GoTo type of internal Link from a rectangular area to the specified pdf.
    This method is different from what I provided in the pdf.py library
    Here you can directly send the page you want to edit instead of the page number
    Helps in doing many things at a time without going through the need of access the file sequentially
    This uses the basic structure of AddUri

    :param int pageId: object id of the page on which to place the URI action.
    :param int uri: string -- uri of resource to link to.
    :param rect: :class:`RectangleObject<PyPDF2.generic.RectangleObject>` or array of four
        integers specifying the clickable rectangular area
        ``[xLL, yLL, xUR, yUR]``, or string in the form ``"[ xLL yLL xUR yUR ]"``.
    :param border: if provided, an array describing border-drawing
        properties. See the PDF spec for details. No border will be
        drawn if this argument is omitted or given [0,0,0].
    -Jayant Solanki
    """
    C = [0,0,1]#border color
    CObject = ArrayObject([NameObject(n) for n in C])
    if border is not None:
        borderArr = [NameObject(n) for n in border[:3]]
        if len(border) == 4:
            dashPattern = ArrayObject([NameObject(n) for n in border[3]])
            borderArr.append(dashPattern)
    else:
        borderArr = [NumberObject(0)] * 3

    if isString(rect):
        rect = NameObject(rect)
    elif isinstance(rect, RectangleObject):
        pass
    else:
        rect = RectangleObject(rect)
    lnk2 = DictionaryObject()
    lnk2.update({
    # NameObject('/NewWindow') : BooleanObject('true'),
    NameObject('/S'): NameObject('/GoTo'),
    NameObject('/D') : ArrayObject([NameObject(pageId), NameObject(0), NameObject('R'), NameObject('/XYZ'), NameObject(0), NameObject(0), NameObject(0)])#inherent zoom, last number should be zero
    });

    lnk = DictionaryObject()
    lnk.update({
    NameObject('/Type'): NameObject('/Annot'),
    NameObject('/Subtype'): NameObject('/Link'),
    # NameObject('/P'): pageLink,
    NameObject('/Rect'): rect,
    NameObject('/H'): NameObject('/I'),
    NameObject('/Border'): ArrayObject(borderArr),
    NameObject('/C'): CObject,
    NameObject('/A'): lnk2
    })
    lnkRef = output._addObject(lnk)# creating new object for indirect reference
    if "/Annots" in page:
        page['/Annots'].append(lnkRef)
    else:
        page[NameObject('/Annots')] = ArrayObject([lnkRef])
    return output, page

def addText(page, text, position, ignoreByteStringObject=False, debug=False):
    """
    Add user given text to a page content
    :page is the page on which text has to be added
    :text is the text string
    :position is the place where it has to be added, examples "top-right", "top-left", "bottom-right", "bottom-left", "bottom-center", "top-center" 
    :param bool ignoreByteStringObject: optional parameter
    to ignore ByteString Objects.
    """
    pageRef = page
    content = pageRef['/Contents'].getObject()
    # if not isinstance(content, ContentStream()):
    content = ContentStream(content, pageRef)#creating contentstream class instance

    # append this line ([], 'BT'), ([1, 0, 0, 1, 52, 34.5], 'Tm'), (['/F1', 9], 'Tf'), ([0, 0, 0], 'rg'), ([u'Study: PCYC-1127-CA'], 'Tj'), ([0], 'g'), ([], 'ET'), ([], 'BT'),
    #  ([1, 0, 0, 1, 336.1, 34.5], 'Tm'), (['/F1', 9], 'Tf'), ([0, 0, 0], 'rg'), ([u'This document is confidential.'], 'Tj'), 
    # ([0], 'g'), ([], 'ET'), ([], 'BT'), ([1, 0, 0, 1, 633.34, 34.5], 'Tm'), (['/F1', 9], 'Tf'), ([0, 0, 0], 'rg'), ([TextStringObject(text)], 'Tj'), ([0], 'g'), ([], 'ET')
    if position == "bottom-center":
        if not ignoreByteStringObject:
            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 52, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([TextStringObject('Study: PCYC-1127-CA')])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))

            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 336.1, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([TextStringObject('This document is confidential.')])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))

            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 633.34, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([TextStringObject(text)])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))

        else:
            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 633.34, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([text])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))


    pageRef.__setitem__(NameObject('/Contents'), content)
    return pageRef

def removeText(page, patterns, altTexts, debug=False, ignoreByteStringObject=False):
    """
    Removes selective text from page's content attribute.

    :param bool ignoreByteStringObject: optional parameter
        to ignore ByteString Objects.
    """
    pageRef = page
    content = pageRef['/Contents'].getObject()
    Content = ContentStream(content, pageRef)#creating contentstream class instance
    # data = Content.operations
    for count in range(0, len(Content.operations)):
        operands, operator = Content.operations[count]
        try:
            for index in range(0, len(patterns)):
                pattern = patterns[index]
                altText = altTexts[index]
                # if operator == b_('rg') and operands == [0,0,1]:#links only
                #   operands[2] = NumberObject(0)
                    # print operands
                if operator == b_('Tj') and re.search(pattern, operands[0], re.M|re.I):
                    text = operands[0]
                    if altText == None:#just strip the link, do not delete the link text
                        replaceString =  text
                        temp, others  = Content.operations[count-1]
                        temp[2] = NumberObject(0)# changing rg array [0,0,1] to [0,0,0]
                    elif altText == '':
                        replaceString =  altText
                        temp, others  = Content.operations[count-3] # 3 obtained by checking the pdf file, can be improved, 
                        temp[0] = TextStringObject(temp[0].replace('(', ''))# deleting the parentheses (, if it is there then it will replace, else wont touch.
                        temp, others  = Content.operations[count+3]
                        temp[0] = TextStringObject(temp[0].replace(')', ''))# deleting the parentheses (
                    elif pattern == r'\bPage \d+$':
                        replaceString = text+altText
                    elif pattern == r'\bProgramming Statements$' and count+10 < len(Content.operations):
                        temp, others  = Content.operations[count+10]
                        if temp[0] == '/F1':
                            temp[1] = NumberObject(9)# changing font size to 9
                        continue

                    else:
                        replaceString = re.sub(pattern, altText, text)# looking for *-Table occurrence in the string #replace a patters with given string
                    if debug:
                        print ("NOTE: *** Replaced string is "+replaceString+" for pattern "+pattern+" ***")
                    if not ignoreByteStringObject:
                        if isinstance(text, TextStringObject):
                            operands[0] = TextStringObject(replaceString)
                    else:
                        if isinstance(text, TextStringObject) or isinstance(text, ByteStringObject):
                            operands[0] = TextStringObject(replaceString)
        except Exception as e:
            print ('WARNING: problem in deciphering the stream page' +str(e))
            print(traceback.format_exc())

    pageRef.__setitem__(NameObject('/Contents'), Content)
    return pageRef


def findLink(input, debug=False):
    processesExtLinks = {}
    with open(extLinks, 'w') as f:
        f.write("Name" + "\t" + "Path" + "\t" + "Type" + "\t" + "ObjectId" + "\t" + "PageNo." + "\t" + "OlderRectCoordinates"+ "\t" + "NewerRectCoordinates"+ "\n")
        for index in range(0,input.getNumPages()):
            # print ("Currently at Page Number: %d" %(index+1))
            page = input.getPage(index)
            # objStm = page['/Contents'].getObject()
            # linkNames = parseContent(objStm.getData())#parsing the contentstream, looking for text with blue formatting
            linkNames = parseContent(page)#parsing the contentstream, looking for text with blue formatting
            key = -1
            if('/Annots' in page):# proceed only if page has links
                for annot in page['/Annots']:
                    try:
                        path = None
                        key = key + 1
                        obj = annot.getObject()
                        if(obj['/A']['/S'] == '/GoToR' or obj['/A']['/S'] == '/Launch' or obj['/A']['/S'] == '/URI'):#look for external links only
                            if debug:
                                print("NOTE: *** Object defining the link is %d ***" %annot.idnum)
                                print("NOTE: *** Link Type: "+str(obj['/A']['/S'])+" ***")
                                print("NOTE: *** Link Coor "+str(obj['/Rect'])+" ***")
                            #sanitise the LinkName here
                            # print(linkNames[index][1:-3])
                            OlderRectCoordinates = obj['/Rect']
                            NewerRectCoordinates = []
                            # matchString = re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I)# looking for -Table occurence in the string
                            if re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I) or re.search(r'^\d*[0-9]-Figure ', linkNames[key], re.M|re.I):
                                findPosDash = linkNames[key].index('-')
                                sanitisedString = linkNames[key][findPosDash+1:] # start after -, 001-Table becomes Table
                                oldWidth = OlderRectCoordinates[3] - OlderRectCoordinates[1]
                                oldWidthperLetter = oldWidth/len(linkNames[key])
                                widthToShorten = (findPosDash-0)*oldWidthperLetter # since 4 characters have been deleted
                                NewerRectCoordinates = OlderRectCoordinates[:]
                                NewerRectCoordinates[3] = NewerRectCoordinates[3] - widthToShorten
                                NewerRectCoordinates[3] = round(NewerRectCoordinates[3],2)# just to two decimal places


                            else:
                                sanitisedString = linkNames[key]
                                NewerRectCoordinates = OlderRectCoordinates[:]
                            if(obj['/A']['/S'] == '/GoToR'):
                                if debug:
                                    if isinstance(obj['/A']['/F'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'])+" ***")
                                if (obj['/A']['/F'] == ''):
                                    path = 'Invalid link'
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                else:
                                    path = obj['/A']['/F']
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : obj['/A']['/D'], 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                f.write(sanitisedString + "\t" + path + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                            if(obj['/A']['/S'] == '/Launch'):
                                if debug:
                                    if isinstance(obj['/A']['/F'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'])+" ***")
                                if (obj['/A']['/F'] == ''):
                                    path = 'Invalid link'
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                else:
                                    path = obj['/A']['/F']
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                f.write(sanitisedString + "\t" + path + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                            if(obj['/A']['/S'] == '/URI'):
                                if debug:
                                    if isinstance(obj['/A']['/URI'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/URI'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/URI'])+" ***")
                                if (obj['/A']['/URI'] == ''):
                                    path = 'Invalid link'
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None,  'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                else:
                                    path = obj['/A']['/URI']
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                f.write(sanitisedString + "\t" + path + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                        else:#for /GoTo, internal links
                            if(obj['/A']['/S'] == '/GoTo'):
                                if debug:
                                    print("NOTE: *** Object defining the link is %d ***" %annot.idnum)
                                    print("NOTE: *** Link Type: "+str(obj['/A']['/S'])+" ***")
                                    print("NOTE: *** Link Coor "+str(obj['/Rect'])+" ***")
                                    if isinstance(obj['/A']['/D'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/D'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/D'])+" ***")
                                OlderRectCoordinates = obj['/Rect']
                                NewerRectCoordinates = OlderRectCoordinates[:]
                                # matchString = re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I)# looking for -Table occurence in the string
                                if re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I) or re.search(r'^\d*[0-9]-Figure ', linkNames[key], re.M|re.I):
                                    findPosDash = linkNames[key].index('-')
                                    sanitisedString = linkNames[key][findPosDash+1:] # start after -, 001-Table becomes Table
                                    oldWidth = OlderRectCoordinates[3] - OlderRectCoordinates[1]
                                    oldWidthperLetter = oldWidth/len(linkNames[key])
                                    widthToShorten = (findPosDash-0)*oldWidthperLetter # since 4 characters have been deleted

                                    NewerRectCoordinates = OlderRectCoordinates[:]
                                    NewerRectCoordinates[3] = NewerRectCoordinates[3] - widthToShorten
                                    NewerRectCoordinates[3] = round(NewerRectCoordinates[3],2)


                                else:
                                    sanitisedString = linkNames[key]
                                    NewerRectCoordinates = OlderRectCoordinates[:]
                                path = obj['/A']['/D'].getObject() # [13, "/XYZ", 0,10000,0]
                                try:# some internal links are invalid
                                    pageNum = input._getPageNumberByIndirect(path[0])# it was not working since the path[0] has indirectobject instead of page number
                                    # path[0] = pageNum # this line disturbs the pointer, which in turn disturbs the pdf writing
                                    f.write(sanitisedString + "\t" + str([pageNum, path[1:-1]]) + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                except:
                                    f.write(sanitisedString + "\t" + 'Invalid link' + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                                    processesExtLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                    except:
                        print ("ERROR: *** Fatal error occurred while looking for links in side the input pdf, exiting now ***")
                        print(traceback.format_exc())
                        progExit()
    return processesExtLinks

def bookMarksRepair(output, outlines, titles, oprMap, debug=False):
    '''
    Bookmark generation
    '''

    parent = None
    subParent = None
    # print outlines
    index = 0
    titleCount = 0
    for item in outlines:
        
        if(isinstance(item, list)):
            for moreitem in item: #check if the item is a nested bookmark
                if(isinstance(moreitem, list)):#super nested # level 3
                    for moremoreitems in moreitem: #check if the item is a nested bookmark
                        other, title = titles[titleCount]
                        if '/Path' in moremoreitems:# check for external link
                            if debug:
                                print ("NOTE: *** Link Path is %s ***" %moremoreitems['/Path'])#object id which encapsulate the page
                                print ("NOTE: *** Bookmark name is "+ moremoreitems['/Title']+" ***")
                                print ("NOTE: *** Alternate Bookmark name is "+ title+" ***")
                            output.addBookmark(title, moremoreitems['/Path'], subParent, None, False, False, '/XYZ', 0, 0 , 0)
                        else:
                            if debug:
                                print ("NOTE: *** Object encapsulating the page is %d ***" %moremoreitems.page.idnum)#object id which encapsulate the page
                                print ("NOTE: *** Bookmark name is "+ moremoreitems['/Title']+" ***")
                                print ("NOTE: *** Alternate Bookmark name is "+ title+" ***")
                            output.addBookmark(title, input._getPageNumberByIndirect(moremoreitems.page.idnum), subParent, None, False, False, '/XYZ', 0, 0 , 0)

                        titleCount = titleCount + 1
                else: # level 2
                    other, title = titles[titleCount]
                    if '/Path' in moreitem:# check for external link
                        if debug:
                            print ("NOTE: *** Link Path is %s ***" %moreitem['/Path'])#object id which encapsulate the page
                            print ("NOTE: *** External Bookmark name is "+ moreitem['/Title']+ " ***")
                            print ("NOTE: *** Alternate Bookmark name is "+title+" ***")
                        if re.search(r'^\d*[0-9]-Table ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Table ', 'Table ', title)# looking for *-Table occurrence in the string
                            subParent = output.addBookmark(replaceString, moreitem['/Path'], parent, None, False, False, '/XYZ', 0, 0 , 0)
                        elif re.search(r'^\d*[0-9]-Figure ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Table ', 'Table ', title)# looking for *-Figure occurrence in the string
                            subParent = output.addBookmark(replaceString, moreitem['/Path'], parent, None, False, False, '/XYZ', 0, 0 , 0)
                        else:
                            subParent = output.addExtBookmark(title, moreitem['/Path'], parent)
                        titleCount = titleCount + 1
                    else:
                        if debug:
                            print ("NOTE: *** Object encapsulating the page is %d ***" %moreitem.page.idnum)#object id which encapsulate the page
                            print ("NOTE: *** Bookmark name is "+ str(moreitem['/Title'])+" ***")
                            print ("NOTE: *** Alternate Bookmark name is "+title+" ***")
                        if re.search(r'^\d*[0-9]-Table ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Table ', 'Table ', title)# looking for *-Table occurrence in the string
                            subParent = output.addBookmark(replaceString, input._getPageNumberByIndirect(moreitem.page.idnum), parent, None, False, False, '/XYZ', 0, 0 , 0)
                        elif re.search(r'^\d*[0-9]-Figure ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Figure ', 'Figure ', title)# looking for *-Figure occurrence in the string
                            subParent = output.addBookmark(replaceString, input._getPageNumberByIndirect(moreitem.page.idnum), parent, None, False, False, '/XYZ', 0, 0 , 0)
                        else:
                            subParent = output.addBookmark(title, input._getPageNumberByIndirect(moreitem.page.idnum), parent, None, False, False, '/XYZ', 0, 0 , 0)
                        titleCount = titleCount + 1

                    
        else: # level 1
            other, title = titles[titleCount]
            if '/Path' in item:# check for external link, had to modify the library to detect external link in the bookmarks
                if debug:
                    print ("NOTE: *** Link Path is %s ***s" %item['/Path'])#object id which encapsulate the page
                    print ("NOTE: *** Bookmark name is "+ str(item['/Title'])+ " ***")
                    print ("NOTE: *** Alternate Bookmark name is "+ title+ " ***")
                if(re.search('Table', title, re.M|re.I)):#omitting table and figures, these are redundant bookmarks
                    titleCount = titleCount + 1
                    continue
                if(re.search('Figure', title, re.M|re.I)):
                    titleCount = titleCount + 1
                    continue
                if (index==1 and  oprMap['Bookmark']['OP002'] and not(re.search(r'Supplemental ', title, re.M|re.I))):#adding the external links, in case it is not present
                    parent = output.addExtBookmark('Supplemental Docs', item['/Path'])
                    output.addExtBookmark('ADRG', "../datasets/ADRG.pdf", parent)
                    output.addExtBookmark('SAP', "../../../misc/PCYC-1127-CA-SAP.pdf", parent)
                parent = output.addBookmark(title, input._getPageNumberByIndirect(item.page.idnum), None, None, False, False, '/XYZ', 0, 0 , 0)
                titleCount = titleCount + 1
            else:
                if debug:
                    print ("NOTE: *** Object encapsulating the page is %d ***" %item.page.idnum)#object id which encapsulate the page
                    print ("NOTE: *** Bookmark name is "+ str(item['/Title'])+ " ***")
                    print ("NOTE: *** Alternate Bookmark name is "+ title+ " ***")
                if(re.search('Table', title, re.M|re.I)):
                    titleCount = titleCount + 1
                    continue
                if(re.search('Figure', title, re.M|re.I)):
                    titleCount = titleCount + 1
                    continue
                if (index==1 and oprMap['Bookmark']['OP002'] and not(re.search(r'Supplemental ', item['/Title'], re.M|re.I))):#adding the external links, in case it is not present
                    parent = output.addBookmark('Supplemental Docs', 0)#point to first page
                    output.addExtBookmark('ADRG', "../datasets/ADRG.pdf", parent)
                    output.addExtBookmark('SAP', "../../../misc/PCYC-1127-CA-SAP.pdf", parent)
                parent = output.addBookmark(title, input._getPageNumberByIndirect(item.page.idnum), None, None, False, False, '/XYZ', 0, 0 , 0)
            index = index + 1
            titleCount = titleCount + 1
    return output


def findPageByText(titleName, subTitle, input, pos=0, pageNo=0, debug=False):
    
    # print content.operations
    tries = 0 # number of pages to look into before breaking the while loop, security check to prevent inifnite loop
    wordsubTitle = ''
    wordflag = False
    pageFound = None
    lineFound =  None# line at which analysis result was found
    while (tries < 3): # can be changed,  # stop looking for the word in the further pages
        page = input.getPage(pageNo)
        extractedText = (page.extractText()).encode('utf-8')
        # print (page.extractText()).encode('utf-8')
        splitSentences = extractedText.split('\n')
        if debug:
            print ("NOTE: *** Looking into Page number "+ str(pageNo)+" ***")
            print ("NOTE: *** Title to search for is "+ titleName+ " ***")
            print ("NOTE: *** Subtitle to search for is "+ subTitle+ " ***")
        count = -1 # new page reset counter
        if(pos != 0):
            flag = True # for identifying the word asccoiated with Table or Figure
        else:
            flag =  False
        for sentence in splitSentences: ## look into the example pdf to understand the logic of this algo
            if (pos>=count): #start processing string after given line number
                count = count + 1
                continue
            try:
                if re.search(titleName, sentence, re.M|re.I): # first parameter, word to be searched, second parameter, word series were word may be present
                    # text = operands[0]
                    flag = True
                    if debug:
                        print ("NOTE: *** Title Found at page"+ str(pageNo)+ " ***")
                    # print ("Word matched is ",operands[0])
                if sentence == 'Analysis Result' and flag: #start building the subsequent words into a full sentence
                    # text = operands[0]
                    # wordflag = True
                    # wordsubTitle = ''
                    pageFound = pageNo # page number where the Analysis Result was found
                    lineFound = count   
                    if debug:
                        print ("NOTE: *** Subtitle Found at page" +str(pageFound) + " ***")
                    return pageFound, lineFound # return the page number at which the string was found
            except:
                # print ('problem in deciphering the stream page')
                pass
            count = count + 1
        pageNo =  pageNo + 1
        tries =  tries + 1
        pos = 0# moved to next page so start searching from first line
    print ("NOTE: *** Not found, going back ***")
    return -1, 0

def fixExtLinks(output, page, index, processesLinks, fileMap, oprMap, debug=False):
    for key in processesLinks:
        if oprMap['ExternalLink']['OP003'] and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/Launch' and processesLinks[key]['Path'] == None: # rule 1, for removing and fixing external links which
            try:
                fileMap[processesLinks[key]['Name']] = fileMap[processesLinks[key]['Name']].replace('_', '-')
                fileMap[processesLinks[key]['Name']] = fileMap[processesLinks[key]['Name']].replace('.sas', '-sas.pdf')
                namepath = '../../../../eSub/misc/'+fileMap[processesLinks[key]['Name']]# this can throw error if the key is incorrect, so I am exploiting it to detect invalid keys
                page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)
                output, page = addGOTOR(output, page, namepath, processesLinks[key]['NewerRectCoordinates'], [0,0,0], 1)
            except Exception as e:
                print ("ERROR: *** In the Excel spec sheet, unable to find entry for "+ str(key) + " ***")
                print(traceback.format_exc())
                progExit()
                # continue
        elif oprMap['TextRemoval']['OP010'] and re.search(r'ad[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
            # print (processesLinks[key]['Name'])
            page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)# deleting the link
            continue
        # elif re.search(r'../../../../eSub/analysis/adam/programs/ad[a-z-A-Z0-9]+-sas.txt$',processesLinks[key]['Name'], re.M|re.I) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
        elif oprMap['TextRemoval']['OP011'] and re.search(r'../[a-z-A-Z0-9]+/ad[a-z-A-Z0-9]+-sas.txt$',processesLinks[key]['Name'], re.M|re.I) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
            # print (processesLinks[key]['Name'])
            page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)# deleting the link
            continue
        elif oprMap['ExternalLink']['OP005'] and (re.search(r't-[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I) or re.search(r'f-[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I) or re.search(r'l-[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I)) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/URI': # rule 2, for removing and fixing external links which have ad*.sas links
            # print (processesLinks[key]['Name'])
            try:
                processesLinks[key]['Path'] = processesLinks[key]['Path'].replace('.txt', '.pdf')
                page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)
                output, page = addGOTOR(output, page, processesLinks[key]['Path'], processesLinks[key]['NewerRectCoordinates'], [0,0,0], 1)
            except Exception as e:
                print ("ERROR: *** while providing file path ***")
                print(traceback.format_exc())
                progExit()
                # continue
        # fixing ADRG section page numbers
        elif oprMap['ExternalLink']['OP004'] and (re.search(r'ADRG Section [0-9]',processesLinks[key]['Name'], re.M|re.I) or re.search(r'SAP Section [0-9]',processesLinks[key]['Name'], re.M|re.I)) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
            try:
                page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)
                if isinstance(processesLinks[key]['Dest'], list): #bug fixing
                    output, page = addGOTOR(output, page, processesLinks[key]['Path'], processesLinks[key]['NewerRectCoordinates'], [0,0,0], processesLinks[key]['Dest'][0])
                else:
                    output, page = addGOTOR(output, page, processesLinks[key]['Path'], processesLinks[key]['NewerRectCoordinates'], [0,0,0], processesLinks[key]['Dest'])
            except Exception as e:
                print ("ERROR: *** while providing file path ***")
                print(traceback.format_exc())
                progExit()
                # continue
            
    return output, page

def fixIntLinks(input, output, page, index, processesLinks, subLinks, debug=False):
    pageNo = -1
    titleName = None
    subTitle = None
    pos = 0
    count = 0
    _pages = output.getObject(output._pages)
    for key in processesLinks: # key is the objectID
        if processesLinks[key]['Type'] == '/GoTo' and processesLinks[key]['Path'] != None and processesLinks[key]['PageNo'] == index:
            pageNo = input._getPageNumberByIndirect(processesLinks[key]['Path'][0])
            titleName = processesLinks[key]['Name']
            pos = 0# no need to reset actually
            count= 0
            # print "Title from the PDF is ", titleName
        elif processesLinks[key]['Type'] == '/GoTo' and processesLinks[key]['Path'] == None and processesLinks[key]['PageNo'] == index: #  and str(processesLinks[key]['Name']): # what did I add the last and operator and why am I checking str(processedLinks), need to verify.
            
            try:
                page = removeIntLinks(page, key)
                if debug:
                    print ("NOTE: *** Current sublink from file is "+subLinks[titleName][count]+" ***")
                    print ("NOTE: *** removed broken link from pdf is "+processesLinks[key]['Name']+ " ***")
                subTitle = processesLinks[key]['Name'] # get it from the tsv file
                sublinks = subLinks[titleName][count].split()[0] # get it from the excel sheet and take the first word
                sublinkToMatch = subTitle.split()[0] # take the first word
                if(sublinks.lower() == sublinkToMatch.lower()):#just matching the first words of subtitle and sublink names
                    newPageNo, newPos = findPageByText(titleName, subTitle, input, pos, pageNo, args.verbose)
                    if count < (len(subLinks[titleName])-1):
                        count =  count + 1;
                    if (newPageNo == -1):
                        print ("WARNING: *** Unable to find exact page number for "+subTitle+" ***")
                        pageId = _pages["/Kids"][pageNo].idnum
                        # output.addLink(index, pageNo, processesLinks[key]['NewerRectCoordinates'], border=None)
                        output, page = addGOTO(output, page, pageId, processesLinks[key]['NewerRectCoordinates'], border=None)# unable to find to as a fail safe point to the page number at whichfigure or table points to
                    else:
                        newpageId = _pages["/Kids"][newPageNo].idnum
                        # output.addLink(index, newPageNo, processesLinks[key]['NewerRectCoordinates'], border=None)
                        output, page = addGOTO(output, page, newpageId, processesLinks[key]['NewerRectCoordinates'], border=None)
                        pageNo = newPageNo # update the last page number, saves some processing
                        pos = newPos
                else:
                    pageId = _pages["/Kids"][pageNo].idnum
                    # output.addLink(index, pageNo, processesLinks[key]['NewerRectCoordinates'], border=None)
                    output, page = addGOTO(output, page, pageId, processesLinks[key]['NewerRectCoordinates'], border=None)
                    if debug:
                        print("NOTE: *** partial link found, linking it to page  "+ str(pageNo)+ " ***")
            except Exception as e:
                print ("ERROR: Exiting, "+str(e))
                print ("ERROR: *** In the Excel spec sheet, unable to find entry for "+ str(titleName)+" ***")
                print(traceback.format_exc())
                progExit()
        
    return output, page



################################################################################################################
#########################  Main program begins  ####################################
###                      parsing the arguments passed                               ###
# setting the log file destination
print ("\n*************************************** pdftoolkit started ***************************************\n")
if args.logFile!="":
    sys.stdout = open(args.logFile, "w")

if(not (re.search(r'.pdf', args.inFile, re.M|re.I) and re.search(r'.pdf', args.outFile, re.M|re.I))):
    print ("ERROR: *** Cannot find .pdf extension ***")
    print ("Please mention the input and output pdf file names in proper format, -h for more info")
    print ("first argument is input file and second argument is output file")
    print ("For example.. ")
    print ("python main.py -i path/input.pdf -o path/output.pdf")
    progExit()

#Checking if the ANASPEC file has been passed by the user
if args.specFile: # if passed then check below for extension
    if(not (re.search(r'.xlsx', args.specFile, re.M|re.I))): #check if the file has proper extension format.
        print ("ERROR: *** Cannot find .xlsx extension for the spec file ***")
        print ("-h or --help for more running the script in proper format")
        print ("Exiting now")
        progExit()

input = PdfFileReader(open(args.inFile, "rb")) # reading the input pdf file
print ("NOTE: *** You are currently running Python PDFTools version 1.0 ***")
sys.stdout.flush()
NumofPages = input.getNumPages()
print ("NOTE: *** %d pages processed Successfully ***"%(NumofPages))
sys.stdout.flush()
print ("NOTE: *** Processing Document "+ args.inFile +" ***")
sys.stdout.flush()
outlines = input.getOutlines() # fetch bookmarks
titles = get_toc(args.inFile) # fetch bookmarks using alternate method, using pdfminer, this gives bug free bookmark names
######################## Creating rules #########################

#checking if the CONFIG file has been passed by the user
if args.configFile: # if passed then check below for extension
    if(not (re.search(r'.xlsx', args.configFile, re.M|re.I))): #check if the file has proper extension format.
        print ("ERROR: *** Cannot find .xlsx extension for the config file ***")
        print ("-h or --help for more running the script in proper format")
        print ("Exiting now")
        progExit()
    print ("NOTE: *** Fetching operation list from the Configuration spreadsheet ***")
    try:
        oprMap = readConfExcel(filename =args.configFile)
    except Exception as e:
        print ('ERROR: *** Some error occurred while reading the config spreadsheet ***')
        print(traceback.format_exc())
else:#if CONFIG file not provided by the user then create the rules/operations yourselves on the basis of ANASPEC file
    #default operations to be performed even if the ANASPEC file is present or not, True means operation to be followed
    oprMap = {'Bookmark': {'OP002': True, 'OP001': True}, 'TextRemoval': {'OP008': False, 'OP009': True, 'OP011': True, 'OP010': True, 'OP012': False, 'OP007': False, 'OP013': True}, 'InternalLink': {'OP006': False}, 'ExternalLink': {'OP003': False, 'OP004': False, 'OP005': False}}
    if checkForAnaspec(outlines):
        if args.specFile: #if ANASPEC file is there then add follwing rules/operations
            #setting following rules for ANASPEC
            oprMap['TextRemoval']['OP007'] =  True
            oprMap['TextRemoval']['OP008'] =  True
            oprMap['TextRemoval']['OP012'] =  True
            oprMap['ExternalLink']['OP003'] =  True
            oprMap['ExternalLink']['OP004'] =  True
            oprMap['ExternalLink']['OP005'] =  True
            oprMap['InternalLink']['OP006'] =  True
        else:
            print ("")
            print ("")
            print ("ERROR: *** The PDF document you provided needs ANASPEC file, please provide this option and try again ***")
            sys.stdout.flush()
            progExit()
    else:
        if args.specFile: #if ANASPEC file is there but is actually not needed
            print ("")
            print ("")
            print ("ERROR: *** The input pdf file does not require an Analysis Result Specification Excel.  Remove this option and try again. ***")
            sys.stdout.flush()
            progExit()


# print ("NOTE: *** "+str(oprMap)+" ***")


# creating TextRemoval rule
patterns = []
altTexts = []
if(oprMap['TextRemoval']['OP007']):
    patterns.append( r'^\d*[0-9]-Table ')
    altTexts.append('Table ')
if(oprMap['TextRemoval']['OP008']):
    patterns.append(r'^\d*[0-9]-Figure ')
    altTexts.append('Figure ')
if(oprMap['TextRemoval']['OP009']):
    patterns.append( r'\bPage \d+$')
    altTexts.append(' of '+str(NumofPages))
if(oprMap['TextRemoval']['OP010']):
    patterns.append(r'\bad[a-z-A-Z0-9]+.sas$')
    altTexts.append(None)
if(oprMap['TextRemoval']['OP011']):
    patterns.append(r'../[a-z-A-Z0-9]+/ad[a-z-A-Z0-9]+-sas.txt$')
    altTexts.append('')
if(oprMap['TextRemoval']['OP012']):
    patterns.append(r'\bProgramming Statements$')
    altTexts.append('Programming Statements')
if(oprMap['TextRemoval']['OP013']):
    patterns.append(r'\bAbbvie$')
    altTexts.append(' ')
#################################################
processesLinks = []
fileMap = []
subLinks = []
################# reading from ANASPEC file ###########################################
if (oprMap['InternalLink']['OP006'] or oprMap['ExternalLink']['OP003']):# only this two opcodes actually needs ANASPEC file, hence if these are active then only read the file
    print ("NOTE: *** Getting TFL mapping from the ANASPEC Excel file ***")
    sys.stdout.flush()
    print ("NOTE: *** TFL mapped Successfully ***")
    sys.stdout.flush()
    fileMap, subLinks = readSpecExcel(filename =args.specFile)# fetch correct filemaping from the given excel file, ANASPEC sheet
    sys.stdout.flush()
################### Fetching all links ################################################
if (oprMap['TextRemoval']['OP010'] or oprMap['TextRemoval']['OP011'] or oprMap['InternalLink']['OP006'] or oprMap['ExternalLink']['OP005'] or oprMap['ExternalLink']['OP004'] or oprMap['ExternalLink']['OP003']):
    print ("NOTE: *** Identifying the all the external and internal links in the PDF doc ***")
    sys.stdout.flush()
    processesLinks = findLink(input, args.verbose)# find all external links from the pdf and store them in a dictionary list for further processing
    print ("NOTE: *** Links identified Successfully ***")
    sys.stdout.flush()
##################################################################################################
#####################################################Processing pages and correcting errors like incomplete links, wrong words etc############################################################

output = PdfFileWriter() ##Opening the output stream
print ("NOTE: *** Performing corrections on the PDF ***")
sys.stdout.flush()
for index in range(0,NumofPages):
        if args.logFile=="" and False:
            sys.stdout.write("\rNOTE: *** Successfully processed: %d of %d pages " %(index+1,NumofPages))
            sys.stdout.flush()
        page = input.getPage(index)
        if(len(patterns)!=0):
            page = removeText(page, patterns, altTexts, args.verbose)# fixing the incorrect Figure names, table names and adding Page numbers
        output, page = fixExtLinks(output, page, index, processesLinks, fileMap, oprMap,  args.verbose)
        page.compressContentStreams() # compress the content stream using the Flat-encode, decreases the size of the document by 3/4
        output.addPage(page)
if oprMap['InternalLink']['OP006']:
    print ("\nNOTE: *** Correcting Internal Links of the PDF ***")
    for index in range(0,NumofPages):
        page = output.getPage(index)            
        output, page = fixIntLinks(input, output, page, index, processesLinks, subLinks, args.verbose)
################################################################################################
if args.logFile!="" and False:
    print ("\nNOTE: *** Corrections applied successfully on all %d pages***"%(NumofPages))
else:
    print ("\nNOTE: *** Corrections applied successfully on all %d pages***"%(NumofPages))
sys.stdout.flush()
print ("NOTE: *** Repairing bookmarks ***")
sys.stdout.flush()
output = bookMarksRepair(output, outlines, titles, oprMap, args.verbose)
print ("NOTE: *** Bookmarks repaired successfully ***")
sys.stdout.flush()
output.setPageMode("/UseOutlines") #This is what tells the PDF to open to bookmarks
# sys.stdout = sys.__stdout__
try:
    os.remove(extLinks)
except:
    pass

print("NOTE: *** Saving the corrected pdf ***")
while True:
    try:
        outputStream = file(args.outFile, "wb")
        break
    except:
        print ("ERROR: *** Corrected pdf cannot be saved as the "+args.outFile+" file is open, please close it and press Enter to continue ***")
        sys.stdout.flush()
        paused =  raw_input("Thanks \n")

output.write(outputStream)
outputStream.close()
print("NOTE: *** Corrected pdf saved successfully ***")

