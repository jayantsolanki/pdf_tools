# -*- coding: utf-8 -*-
"""
****************************************************************************************************
*                  
*                          ================================
*  This software is intended for correcting deviations introduced in the Define.pdf
*  Please go through the documentation of PYPDF2 get a general idea
****************************************************************************************************
*
*  Project          :  Statistical Computing Platform
*  filename         :  pdftool.py
*  version          :  1.0
*  Study            :  Any
*  Author           :  Jayant Solanki
*  Maintainer       :  
*  Creator          :  Jayant Solanki
*  Date             :  07/23/2018
*  Updated          :  09/16/2018
*  Note             :  This is program which can be included in any other SAS script using %fixpdf macro
*  Description      :  To fix the deviations introduced in the Define.pdf
*  Documentation 1  :  https://pythonhosted.org/PyPDF2/
*  Documentation 2  :  https://github.com/mstamy2/PyPDF2 (Github)
*
****************************************************************************************************
************************************* Necessary libraries ******************************************
Requirements
3.1.    Insert Page X of Y                      : PYPDF2 
3.2.    Bookmark Parent Child                   : PYPDF2
3.3.    Remove Prefixes                         : PYPDF2
3.4.    ADRG and SAP                            : PYPDF2
3.5.    Relative External Links                 : PYPDF2, openpyxl
3.6.    Analysis Results TOC Internal Links     : PYPDF2, openpyxl
3.7.    Retain Initial View                     : PYPDF2
3.8.    Table External Links                    : PYPDF2, openpyxl
3.9.    Remove ADaM Programs (ad*.sas) Links    : PYPDF2
3.10.   Update TFL Program External Links       : PYPDF2, openpyxl
3.11.   Body ADRG and SAP Links                 : PYPDF2
3.14.   Program Code Fonts                      : PYPDF2
3.15.   Remove “AbbVie” from Title Page         : PYPDF2
                        ================================
Miscellaneous
Parsing user provided command-line arguments     : argparse
Reading Excel file                              : openpyxl
Reading Bookmarks                               : pdfminer
Generating random string                        : uuid
Creating regex patterns                         : re
Error report                                    : traceback
Handling system files                           : os
Redirecting debug/text outputs of program       : sys
Exiting program in case of fatal error          : sys
****************************************************************************************************
"""

from PyPDF2 import PdfFileWriter, PdfFileReader # importing class for reading Define.pdf and Writing modified Define.pdf
from PyPDF2.pdf import ContentStream # importing class for reading /Content of each page
from PyPDF2.generic import NameObject, NumberObject, ArrayObject, TextStringObject, RectangleObject, DictionaryObject, BooleanObject, IndirectObject
from PyPDF2.utils import isString, b_
from openpyxl import load_workbook # library function for reading the excel file
import re # regex patterns
import sys # for command line arguments
from argparse import ArgumentParser
import traceback # detailed error report, such as crash due syntax error
# this library has been added for only to fetch the bookmarks titles
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import os
import uuid

"""
************************************ End of libraries **********************************************
"""

sys.setrecursionlimit(10000) # 10000 is an example, try with different values if current number doesn't suit, it tells about how much will be recursion level

# setting up the command-line parameter structure for the python script
parser = ArgumentParser(description='Python based toolkit for repairing define.pdf')
parser.add_argument("-c", "--configFile", help="Takes the path for the config spreadsheet", required=False)#optional
parser.add_argument("-s", "--specFile", help="Takes the specification excel file", default = "", required=False)#optional
parser.add_argument("-i", "--inFile", help="Takes input pdf file for modification", required=True)#required
parser.add_argument("-o", "--outFile", help="Writes the pdf file with the given name", default = "", required=False)#optional
parser.add_argument("-l", "--logFile", help="Writes log messages on a file", default = "", required=False)#optional
parser.add_argument("-v", "--verbose", help="Enable verbose python log", default = False, required=False)#optional
args = parser.parse_args()
extLinks = str(uuid.uuid4())+".tsv" #name for the spreadsheet which will contain the links identified in the Define.pdf document
print args
print args['inFile']
sys.exit()
"""
*********************************** Important Methods **********************************************
****************************************************************************************************
Requirements
3.1.    Insert Page X of Y                      : removeText
3.2.    Bookmark Parent Child                   : get_toc, bookMarksRepair
3.3.    Remove Prefixes                         : removeText
3.4.    ADRG and SAP                            : get_toc, bookMarksRepair, PyPDF2.pdf.addExtBookmark (I added it inside in pdf.py)
3.5.    Relative External Links                 : fixExtLinks, addGOTOR, removeExtLinks 
3.6.    Analysis Results TOC Internal Links     : fixIntLinks, addGOTO, removeIntLinks, findPageByText 
3.7.    Retain Initial View                     : addGOTOR
3.8.    Table External Links                    : fixExtLinks, addGOTOR, removeExtLinks
3.9.    Remove ADaM Programs (ad*.sas) Links    : removeText, removeExtLinks
3.10.   Update TFL Program External Links       : fixExtLinks, addGOTOR, removeExtLinks
3.11.   Body ADRG and SAP Links                 : fixExtLinks, addGOTOR, removeExtLinks
3.14.   Program Code Fonts                      : removeText
3.15.   Remove “AbbVie” from Title Page         : removeText
                        ================================
Miscellaneous
Finding all the links present in pdf            : findLink, parseContent
Mapping TFL names with filepath                 : readSpecExcel
Mapping TFL names with sub-headings             : readSpecExcel
Reading developer defined opcodes               : readConfExcel
Safely exit program                          : progExit
****************************************************************************************************
************************************** End of Methods **********************************************
"""

"""
************************************* Method: addGOTO **********************************************
Requirements:
    3.6.    Analysis Results TOC Internal Links
Description:
    Add an GoTo internal Link from a rectangular area to the specified pdf.
    This method is different from what is provided in the pdf.py library
    Here you can directly send the page you want to edit instead of the page number
    Helps in doing many things at a time without going through the need of access the file sequentially
    This uses the basic structure of AddUri

:param PDFFileWriter instance output: for storing the any new object created during link addition
:param PageObject page: object representation of pdf's page
:param int pageId: object Id of the destination page
:param rect: :class:`RectangleObject<PyPDF2.generic.RectangleObject>` or array of four
    integers specifying the clickable rectangular area
    ``[xLL, yLL, xUR, yUR]``, or string in the form ``"[ xLL yLL xUR yUR ]"``
:param border: if provided, an array describing border-drawing
    properties. See the PDF spec for details. No border will be
    drawn if this argument is omitted or given [0,0,0]
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return FilePDFWriter output, PageObject page: returns the modified file output stream and the modified page
"""
def addGOTO(output, page, pageId, rect, border=None, debug=False):

    C = [0,0,1]#border color
    CObject = ArrayObject([NameObject(n) for n in C])
    if border is not None:
        borderArr = [NameObject(n) for n in border[:3]]
        if len(border) == 4:
            dashPattern = ArrayObject([NameObject(n) for n in border[3]])
            borderArr.append(dashPattern)
    else:
        borderArr = [NumberObject(0)] * 3

    if isString(rect):
        rect = NameObject(rect)
    elif isinstance(rect, RectangleObject):
        pass
    else:
        rect = RectangleObject(rect)
    lnk2 = DictionaryObject()
    lnk2.update({
    NameObject('/S'): NameObject('/GoTo'),
    NameObject('/D') : ArrayObject([NameObject(pageId), NameObject(0), NameObject('R'), NameObject('/XYZ'), NameObject(0), NameObject(0), NameObject(0)])#inherent zoom, last number should be zero
    });

    lnk = DictionaryObject()
    lnk.update({
    NameObject('/Type'): NameObject('/Annot'),
    NameObject('/Subtype'): NameObject('/Link'),
    NameObject('/Rect'): rect,
    NameObject('/H'): NameObject('/I'),
    NameObject('/Border'): ArrayObject(borderArr),
    NameObject('/C'): CObject,
    NameObject('/A'): lnk2
    })
    lnkRef = output._addObject(lnk)# creating new object for indirect reference
    if "/Annots" in page:
        page['/Annots'].append(lnkRef)
    else:
        page[NameObject('/Annots')] = ArrayObject([lnkRef])
    return output, page

"""
************************************* Method: addGOTOR *********************************************
Requirements: 
    3.5.    Relative External Links
    3.7.    Retain Initial View  
    3.8.    Table External Links                 
    3.9.    Remove ADaM Programs (ad*.sas) Links  
    3.10.   Update TFL Program External Links     
    3.11.   Body ADRG and SAP Links 
Description:
    Add an GoToR type of external Link from a rectangular area to the specified pdf.
    This method is different from what is provided in the pdf.py library
    Here you can directly send the page you want to edit instead of the page number
    Helps in doing many things at a time without going through the need of access the file sequentially
    This uses the basic structure of AddUri

:param PDFFileWriter instance output: for storing the any new object created during link addition
:param PageObject page: object representation of pdf's page
:param string path: for storing the external file-path for link
:param rect: :class:`RectangleObject<PyPDF2.generic.RectangleObject>` or array of four
    integers specifying the clickable rectangular area
    ``[xLL, yLL, xUR, yUR]``, or string in the form ``"[ xLL yLL xUR yUR ]"``
:param border: if provided, an array describing border-drawing
    properties. See the PDF spec for details. No border will be
    drawn if this argument is omitted or given [0,0,0]
:param int destination: stores the destination page number which will the external pdf file will open at
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return FilePDFWriter output, PageObject page: returns the modified file output stream and the modified page
"""
def addGOTOR(output, page, path, rect, border=None, destination = 0, debug=False):

    C = [0,0,1]#border color
    CObject = ArrayObject([NameObject(n) for n in C])
    if border is not None:
        borderArr = [NameObject(n) for n in border[:3]]
        if len(border) == 4:
            dashPattern = ArrayObject([NameObject(n) for n in border[3]])
            borderArr.append(dashPattern)
    else:
        borderArr = [NumberObject(0)] * 3

    if isString(rect):
        rect = NameObject(rect)
    elif isinstance(rect, RectangleObject):
        pass
    else:
        rect = RectangleObject(rect)
    lnk2 = DictionaryObject()
    lnk2.update({
    NameObject('/S'): NameObject('/GoToR'),
    NameObject('/F'): TextStringObject(path), # had to do like this other wise not working
    NameObject('/NewWindow true /D') : ArrayObject([NumberObject(int(destination)-1), NameObject('/XYZ'), NameObject(0), NameObject(10000), NameObject(0.0)])#inherent zoom, last number should be zero
    });
    lnk = DictionaryObject()
    lnk.update({
    NameObject('/Type'): NameObject('/Filespec'),
    NameObject('/Subtype'): NameObject('/Link'),
    NameObject('/Rect'): rect,
    NameObject('/H'): NameObject('/I'),
    NameObject('/Border'): ArrayObject(borderArr),
    NameObject('/C'): CObject,
    NameObject('/A'): lnk2
    })
    lnkRef = output._addObject(lnk)# creating new object for indirect reference
    if "/Annots" in page:
        page['/Annots'].append(lnkRef)
    else:
        page[NameObject('/Annots')] = ArrayObject([lnkRef])
    return output, page

"""
************************************* Method: addLAUNCH ********************************************
Requirements: 
    Not used right now
Description:
    Add an Launch type of external Link from a rectangular area to the specified pdf.
    This method is different from what is provided in the pdf.py library
    Here you can directly send the page you want to edit instead of the page number
    Helps in doing many things at a time without going through the need of access the file sequentially
    This uses the basic structure of AddUri

:param PDFFileWriter instance output: for storing the any new object created during link addition
:param PageObject page: object representation of pdf's page
:param string path: for storing the external file-path for link
:param rect: :class:`RectangleObject<PyPDF2.generic.RectangleObject>` or array of four
    integers specifying the clickable rectangular area
    ``[xLL, yLL, xUR, yUR]``, or string in the form ``"[ xLL yLL xUR yUR ]"``
:param border: if provided, an array describing border-drawing
    properties. See the PDF spec for details. No border will be
    drawn if this argument is omitted or given [0,0,0]
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return FilePDFWriter output, PageObject page: returns the modified file output stream and the modified page
"""
def addLAUNCH(output, page, path, rect, border=None, debug=False):

    C = [0,0,1]#border color
    CObject = ArrayObject([NameObject(n) for n in C])
    if border is not None:
        borderArr = [NameObject(n) for n in border[:3]]
        if len(border) == 4:
            dashPattern = ArrayObject([NameObject(n) for n in border[3]])
            borderArr.append(dashPattern)
    else:
        borderArr = [NumberObject(0)] * 3

    if isString(rect):
        rect = NameObject(rect)
    elif isinstance(rect, RectangleObject):
        pass
    else:
        rect = RectangleObject(rect)
    lnk2 = DictionaryObject()
    lnk2.update({
    NameObject('/S'): NameObject('/Launch'),
    NameObject('/F'): TextStringObject(path),
    NameObject('/NewWindow') : BooleanObject('true'),
    });
    lnk = DictionaryObject()
    lnk.update({
    NameObject('/Type'): NameObject('/Annot'),
    NameObject('/Subtype'): NameObject('/Link'),
    NameObject('/Rect'): rect,
    NameObject('/H'): NameObject('/I'),
    NameObject('/Border'): ArrayObject(borderArr),
    NameObject('/C'): CObject,
    NameObject('/A'): lnk2
    })
    lnkRef = output._addObject(lnk)# creating new object for indirect reference
    if "/Annots" in page:
        page['/Annots'].append(lnkRef)
    else:
        page[NameObject('/Annots')] = ArrayObject([lnkRef])
    return output, page

"""
************************************* Method: addText *********************************************
Requirements: 
    Not used right now
Description:
    Add user provided text to a page content
:param PageObject page: is the page on which text has to be added
:param string text: is the text string
:param string position: is the place where it has to be added, examples "top-right", "top-left", "bottom-right", "bottom-left", "bottom-center", "top-center" 
:param bool ignoreByteStringObject: optional parameter
to ignore ByteString Objects.
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return PageObject pageRef: returns the modified page
"""
def addText(page, text, position, ignoreByteStringObject=False, debug=False):

    pageRef = page
    content = pageRef['/Contents'].getObject()
    # if not isinstance(content, ContentStream()):
    content = ContentStream(content, pageRef)#creating Contentstream class instance

    # append this line ([], 'BT'), ([1, 0, 0, 1, 52, 34.5], 'Tm'), (['/F1', 9], 'Tf'), ([0, 0, 0], 'rg'), ([u'Study: PCYC-1127-CA'], 'Tj'), ([0], 'g'), ([], 'ET'), ([], 'BT'),
    #  ([1, 0, 0, 1, 336.1, 34.5], 'Tm'), (['/F1', 9], 'Tf'), ([0, 0, 0], 'rg'), ([u'This document is confidential.'], 'Tj'), 
    # ([0], 'g'), ([], 'ET'), ([], 'BT'), ([1, 0, 0, 1, 633.34, 34.5], 'Tm'), (['/F1', 9], 'Tf'), ([0, 0, 0], 'rg'), ([TextStringObject(text)], 'Tj'), ([0], 'g'), ([], 'ET')
    if position == "bottom-center":
        if not ignoreByteStringObject:
            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 52, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([TextStringObject('Study: PCYC-1127-CA')])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))

            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 336.1, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([TextStringObject('This document is confidential.')])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))

            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 633.34, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([TextStringObject(text)])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))

        else:
            operands = ArrayObject([])
            operator = NameObject("BT") 
            content.operations.append((operands, operator))
            operands = ArrayObject([1, 0, 0, 1, 633.34, 34.5])
            operator = NameObject("Tm") 
            content.operations.append((operands, operator))
            operands = ArrayObject(['/F1', 9])
            operator = NameObject("Tf") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0, 0, 0])
            operator = NameObject("rg") 
            content.operations.append((operands, operator))
            operands = ArrayObject([text])
            operator = NameObject("Tj") 
            content.operations.append((operands, operator))
            operands = ArrayObject([0])
            operator = NameObject("g")  
            content.operations.append((operands, operator))
            operands = ArrayObject([])
            operator = NameObject("ET") 
            content.operations.append((operands, operator))


    pageRef.__setitem__(NameObject('/Contents'), content)
    return pageRef


"""
************************************* Method: bookMarksRepair *************************************
Requirements: 
    3.2.    Bookmark Parent Child  
    3.4.    ADRG and SAP 
Description:
    Regenerates the Bookmarks, removes bad bookmark children and prefixes and also adds external bookmarks inside Supplemental Docs
:param PDFFileWriter instance output: for storing the any new object created during link addition
:param Dictionary list outlines: stores the bookmarks fetched using PYPDF2
:param Dictionary list titles: stores the bookmarks fetched using PDFMiner
:param Dictionary list oprMap: stores opCodes describing kind of operations to perform, refer to config.xlsx for more description
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return FilePDFWriter output: returns the modified file output stream
"""
def bookMarksRepair(output, outlines, titles, oprMap, debug=False):


    parent = None
    subParent = None
    # print outlines
    index = 0
    titleCount = 0
    for item in outlines:
        
        if(isinstance(item, list)):
            for moreitem in item: #check if the item is a nested bookmark
                if(isinstance(moreitem, list)):#super nested # level 3
                    for moremoreitems in moreitem: #check if the item is a nested bookmark
                        other, title = titles[titleCount]
                        if '/Path' in moremoreitems:# check for external link
                            if debug:
                                print ("NOTE: *** Link Path is %s ***" %moremoreitems['/Path'])#object id which encapsulate the page
                                print ("NOTE: *** Bookmark name is "+ moremoreitems['/Title']+" ***")
                                print ("NOTE: *** Alternate Bookmark name is "+ title+" ***")
                            output.addBookmark(title, moremoreitems['/Path'], subParent, None, False, False, '/XYZ', 0, 0 , 0)
                        else:
                            if debug:
                                print ("NOTE: *** Object encapsulating the page is %d ***" %moremoreitems.page.idnum)#object id which encapsulate the page
                                print ("NOTE: *** Bookmark name is "+ moremoreitems['/Title']+" ***")
                                print ("NOTE: *** Alternate Bookmark name is "+ title+" ***")
                            output.addBookmark(title, input._getPageNumberByIndirect(moremoreitems.page.idnum), subParent, None, False, False, '/XYZ', 0, 0 , 0)

                        titleCount = titleCount + 1
                else: # level 2
                    other, title = titles[titleCount]
                    if '/Path' in moreitem:# check for external link
                        if debug:
                            print ("NOTE: *** Link Path is %s ***" %moreitem['/Path'])#object id which encapsulate the page
                            print ("NOTE: *** External Bookmark name is "+ moreitem['/Title']+ " ***")
                            print ("NOTE: *** Alternate Bookmark name is "+title+" ***")
                        if re.search(r'^\d*[0-9]-Table ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Table ', 'Table ', title)# looking for *-Table occurrence in the string
                            subParent = output.addBookmark(replaceString, moreitem['/Path'], parent, None, False, False, '/XYZ', 0, 0 , 0)
                        elif re.search(r'^\d*[0-9]-Figure ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Table ', 'Table ', title)# looking for *-Figure occurrence in the string
                            subParent = output.addBookmark(replaceString, moreitem['/Path'], parent, None, False, False, '/XYZ', 0, 0 , 0)
                        else:
                            subParent = output.addExtBookmark(title, moreitem['/Path'], parent)
                        titleCount = titleCount + 1
                    else:
                        if debug:
                            print ("NOTE: *** Object encapsulating the page is %d ***" %moreitem.page.idnum)#object id which encapsulate the page
                            print ("NOTE: *** Bookmark name is "+ str(moreitem['/Title'])+" ***")
                            print ("NOTE: *** Alternate Bookmark name is "+title+" ***")
                        if re.search(r'^\d*[0-9]-Table ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Table ', 'Table ', title)# looking for *-Table occurrence in the string
                            subParent = output.addBookmark(replaceString, input._getPageNumberByIndirect(moreitem.page.idnum), parent, None, False, False, '/XYZ', 0, 0 , 0)
                        elif re.search(r'^\d*[0-9]-Figure ', title, re.M|re.I):
                            replaceString = re.sub(r'^\d*[0-9]-Figure ', 'Figure ', title)# looking for *-Figure occurrence in the string
                            subParent = output.addBookmark(replaceString, input._getPageNumberByIndirect(moreitem.page.idnum), parent, None, False, False, '/XYZ', 0, 0 , 0)
                        else:
                            subParent = output.addBookmark(title, input._getPageNumberByIndirect(moreitem.page.idnum), parent, None, False, False, '/XYZ', 0, 0 , 0)
                        titleCount = titleCount + 1

                    
        else: # level 1
            other, title = titles[titleCount]
            if '/Path' in item:# check for external link, had to modify the library to detect external link in the bookmarks
                if debug:
                    print ("NOTE: *** Link Path is %s ***s" %item['/Path'])#object id which encapsulate the page
                    print ("NOTE: *** Bookmark name is "+ str(item['/Title'])+ " ***")
                    print ("NOTE: *** Alternate Bookmark name is "+ title+ " ***")
                if(re.search('Table', title, re.M|re.I)):#omitting table and figures, these are redundant bookmarks
                    titleCount = titleCount + 1
                    continue
                if(re.search('Figure', title, re.M|re.I)):
                    titleCount = titleCount + 1
                    continue
                if (index==1 and  oprMap['Bookmark']['OP002'] and not(re.search(r'Supplemental ', title, re.M|re.I))):#adding the external links, in case it is not present
                    parent = output.addExtBookmark('Supplemental Docs', item['/Path'])
                    output.addExtBookmark('ADRG', "../datasets/ADRG.pdf", parent)
                    output.addExtBookmark('SAP', "../../../misc/PCYC-1127-CA-SAP.pdf", parent)
                parent = output.addBookmark(title, input._getPageNumberByIndirect(item.page.idnum), None, None, False, False, '/XYZ', 0, 0 , 0)
                titleCount = titleCount + 1
            else:
                if debug:
                    print ("NOTE: *** Object encapsulating the page is %d ***" %item.page.idnum)#object id which encapsulate the page
                    print ("NOTE: *** Bookmark name is "+ str(item['/Title'])+ " ***")
                    print ("NOTE: *** Alternate Bookmark name is "+ title+ " ***")
                if(re.search('Table', title, re.M|re.I)):
                    titleCount = titleCount + 1
                    continue
                if(re.search('Figure', title, re.M|re.I)):
                    titleCount = titleCount + 1
                    continue
                if (index==1 and oprMap['Bookmark']['OP002'] and not(re.search(r'Supplemental ', item['/Title'], re.M|re.I))):#adding the external links, in case it is not present
                    parent = output.addBookmark('Supplemental Docs', 0)#point to first page
                    output.addExtBookmark('ADRG', "../datasets/ADRG.pdf", parent)
                    output.addExtBookmark('SAP', "../../../misc/PCYC-1127-CA-SAP.pdf", parent)
                parent = output.addBookmark(title, input._getPageNumberByIndirect(item.page.idnum), None, None, False, False, '/XYZ', 0, 0 , 0)
            index = index + 1
            titleCount = titleCount + 1
    return output

"""
************************************* Method: checkForAnaspec *************************************
Requirements: 
    General
Description:
    This method looks for the presence of the "Analysis Result Metadata" in order to make sure
    that the documents require ANASPEC excel file or not
:param Dictionary List outlines: Dictionary list containing all the bookmarks discovered in the pdf
:returns boolean: True or False 
"""

def checkForAnaspec(outlines):

    for item in outlines:
        if(isinstance(item, list)):
            continue
        if (re.search(r'Analysis Results Metadata', str(item['/Title']), re.M|re.I)):
            return True
    return False

"""
************************************* Method: findLink ********************************************
Requirements: 
    General
Description:
    Most important method in the pdftools.py,
    In first step it calls parseContent method to get list of all the blue-colored text strings (links are in blue) page by page basis
    In second step it fetches /Annots page by page and maps those blue-colored text strings with corresponding link's object id
    In third step, it fetches all the important properties of each link in the /Annots and start storing them in processLinks dictionary list
:param PDFFileReader instance input: variable storing the input pdf stream
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return Dictionary List processesLinks: returns the links discovered in the pdf with objectId as key
"""

def findLink(input, debug=False):

    processesLinks = {}
    with open(extLinks, 'w') as f:
        f.write("Name" + "\t" + "Path" + "\t" + "Type" + "\t" + "ObjectId" + "\t" + "PageNo." + "\t" + "OlderRectCoordinates"+ "\t" + "NewerRectCoordinates"+ "\n")
        for index in range(0,input.getNumPages()):
            page = input.getPage(index)
            linkNames = parseContent(page)#parsing the Contentstream, looking for text with blue formatting
            key = -1
            if('/Annots' in page):# proceed only if page has links
                for annot in page['/Annots']:
                    try:
                        path = None
                        key = key + 1
                        obj = annot.getObject()
                        if(obj['/A']['/S'] == '/GoToR' or obj['/A']['/S'] == '/Launch' or obj['/A']['/S'] == '/URI'):#look for external links only
                            if debug:
                                print("NOTE: *** Object defining the link is %d ***" %annot.idnum)
                                print("NOTE: *** Link Type: "+str(obj['/A']['/S'])+" ***")
                                print("NOTE: *** Link Coor "+str(obj['/Rect'])+" ***")
                            #sanitize the LinkName here
                            OlderRectCoordinates = obj['/Rect']
                            NewerRectCoordinates = []
                            if re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I) or re.search(r'^\d*[0-9]-Figure ', linkNames[key], re.M|re.I):
                                findPosDash = linkNames[key].index('-')
                                sanitisedString = linkNames[key][findPosDash+1:] # start after -, 001-Table becomes Table
                                oldWidth = OlderRectCoordinates[3] - OlderRectCoordinates[1]
                                oldWidthperLetter = oldWidth/len(linkNames[key])
                                widthToShorten = (findPosDash-0)*oldWidthperLetter # since 4 characters have been deleted
                                NewerRectCoordinates = OlderRectCoordinates[:]
                                NewerRectCoordinates[3] = NewerRectCoordinates[3] - widthToShorten
                                NewerRectCoordinates[3] = round(NewerRectCoordinates[3],2)# just to two decimal places

                            else:
                                sanitisedString = linkNames[key]
                                NewerRectCoordinates = OlderRectCoordinates[:]
                            if(obj['/A']['/S'] == '/GoToR'):
                                if debug:
                                    if isinstance(obj['/A']['/F'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'])+" ***")
                                if (obj['/A']['/F'] == ''):
                                    path = 'Invalid link'
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                else:
                                    path = obj['/A']['/F']
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : obj['/A']['/D'], 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                f.write(sanitisedString + "\t" + path + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                            if(obj['/A']['/S'] == '/Launch'):
                                if debug:
                                    if isinstance(obj['/A']['/F'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/F'])+" ***")
                                if (obj['/A']['/F'] == ''):
                                    path = 'Invalid link'
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                else:
                                    path = obj['/A']['/F']
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                f.write(sanitisedString + "\t" + path + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                            if(obj['/A']['/S'] == '/URI'):
                                if debug:
                                    if isinstance(obj['/A']['/URI'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/URI'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/URI'])+" ***")
                                if (obj['/A']['/URI'] == ''):
                                    path = 'Invalid link'
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None,  'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                else:
                                    path = obj['/A']['/URI']
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                f.write(sanitisedString + "\t" + path + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                        else:#for /GoTo, internal links
                            if(obj['/A']['/S'] == '/GoTo'):
                                if debug:
                                    print("NOTE: *** Object defining the link is %d ***" %annot.idnum)
                                    print("NOTE: *** Link Type: "+str(obj['/A']['/S'])+" ***")
                                    print("NOTE: *** Link Coor "+str(obj['/Rect'])+" ***")
                                    if isinstance(obj['/A']['/D'], IndirectObject):
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/D'].getObject())+" ***")
                                    else:
                                        print ("NOTE: *** Link Path is "+str(obj['/A']['/D'])+" ***")
                                OlderRectCoordinates = obj['/Rect']
                                NewerRectCoordinates = OlderRectCoordinates[:]
                                # matchString = re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I)# looking for -Table occurrence in the string
                                if re.search(r'^\d*[0-9]-Table ', linkNames[key], re.M|re.I) or re.search(r'^\d*[0-9]-Figure ', linkNames[key], re.M|re.I):
                                    findPosDash = linkNames[key].index('-')
                                    sanitisedString = linkNames[key][findPosDash+1:] # start after -, 001-Table becomes Table
                                    oldWidth = OlderRectCoordinates[3] - OlderRectCoordinates[1]
                                    oldWidthperLetter = oldWidth/len(linkNames[key])
                                    widthToShorten = (findPosDash-0)*oldWidthperLetter # since 4 characters have been deleted

                                    NewerRectCoordinates = OlderRectCoordinates[:]
                                    NewerRectCoordinates[3] = NewerRectCoordinates[3] - widthToShorten
                                    NewerRectCoordinates[3] = round(NewerRectCoordinates[3],2)


                                else:
                                    sanitisedString = linkNames[key]
                                    NewerRectCoordinates = OlderRectCoordinates[:]
                                path = obj['/A']['/D'].getObject() # [13 0 R, "/XYZ", 0,10000,0]
                                try:# some internal links are invalid
                                    pageNum = input._getPageNumberByIndirect(path[0])# it was not working since the path[0] has indirectobject instead of page number
                                    # path[0] = pageNum # this line disturbs the pointer, which in turn disturbs the pdf writing
                                    f.write(sanitisedString + "\t" + str([pageNum, path[1:-1]]) + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : path, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                                except:
                                    f.write(sanitisedString + "\t" + 'Invalid link' + "\t" + obj['/A']['/S'] + "\t" + str(annot.idnum) + "\t" + str(index+1) + "\t" + str(OlderRectCoordinates) + "\t" + str(NewerRectCoordinates) + "\n")
                                    processesLinks[annot.idnum] = {'Name' : sanitisedString , 'Path' : None, 'Dest' : None, 'Type' : obj['/A']['/S'], 'ObjectId' : annot.idnum, 'PageNo' : index , 'OlderRectCoordinates' : OlderRectCoordinates, 'NewerRectCoordinates' : NewerRectCoordinates}
                    except:
                        print ("ERROR: *** Fatal error occurred while looking for links in side the input pdf, exiting now ***")
                        print(traceback.format_exc())
                        progExit()
    return processesLinks

"""
************************************* Method: findPageByText **************************************
Requirements: 
    3.6.    Analysis Results TOC Internal Links
Description:
    Helper method to fixIntLinks method, it is a little complicated algorithm
    It identifies all the sub headings for the TFL data
    It does by first searching for Title name (TFL name), once it is found it start looking for Subsequent Analysis Results which have sub-headings for the Title
    It returns page number for each sub-heading and that is used by fixIntLinks to fix the internal links under Analysis Results Metadata Page
:param string titleName: stores the TFL name such as Table 14.1.2.4 or Figure 12.5.2.3
:param string subTitle: stores the sub-heading which falls under TFL's Analysis Results Metadata, such as Hazard Ratio, P-Value etc
:param PDFFileReader instance input: PDF input stream
:param int pos: stores the position line number from where to start looking for string in the Contentstream
:param int pageNo: stores the page number from where to start looking from
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return int pageNo, int pos: returns the page number and the line number where sub-heading was found, in case not found, return -1 and 0
"""

def findPageByText(titleName, subTitle, input, pos=0, pageNo=0, debug=False):

    tries = 0 # number of pages to look into before breaking the while loop, security check to prevent infinite loop
    pageFound = None
    lineFound =  None # line at which analysis result heading was found
    while (tries < 3): # can be changed,  # stop looking for the word in the further pages
        page = input.getPage(pageNo)
        extractedText = (page.extractText()).encode('utf-8') # to support Unicode format, otherwise some of the text in page won't be recognizable
        splitSentences = extractedText.split('\n')
        if debug:
            print ("NOTE: *** Looking into Page number "+ str(pageNo)+" ***")
            print ("NOTE: *** Title to search for is "+ titleName+ " ***")
            print ("NOTE: *** Subtitle to search for is "+ subTitle+ " ***")
        count = -1 # for every new page reset the line counter
        if(pos != 0):
            flag = True # for identifying the word associated with Table or Figure, if set to true then directly search for "Analysis Result" string, else find Table/Figure Name first
        else:
            flag =  False
        for sentence in splitSentences: ## look into the example pdf to understand the logic of this algo
            if (pos>=count): #start processing string after given line number
                count = count + 1
                continue
            try:
                if re.search(titleName, sentence, re.M|re.I): # look for TFl name, first parameter: word to be searched, second parameter: word series in which word may be present
                    flag = True
                    if debug:
                        print ("NOTE: *** Title Found at page"+ str(pageNo)+ " ***")
                if sentence == 'Analysis Result' and flag: #start looking for Analysis Result, because it has the sub-heading for the Parent TFL name, there can be any number of Analysis Result for a particular TFL name
                    pageFound = pageNo # page number where the Analysis Result was found
                    lineFound = count   
                    if debug:
                        print ("NOTE: *** Subtitle Found at page" +str(pageFound) + " ***")
                    return pageFound, lineFound # return the page number and the line number at which the string or sub-heading was was found
            except:
                pass
            count = count + 1
        pageNo =  pageNo + 1
        tries =  tries + 1
        pos = 0# moved to next page so start searching from first line
    print ("NOTE: *** Not found, going back ***")
    return -1, 0

"""
************************************* Method: fixExtLinks *****************************************
Requirements: 
    3.5.    Relative External Links
    3.7.    Retain Initial View  
    3.8.    Table External Links                 
    3.9.    Remove ADaM Programs (ad*.sas) Links  
    3.10.   Update TFL Program External Links     
    3.11.   Body ADRG and SAP Links 
Description:
    Method to fix the External Links deviation in the pdf
    First, go through the processesLinks to fetch the object ID for each external links which are broken
    Second, delete those existing links and create new links with file paths mentioned
:param PDFFileWriter instance output: PDF output stream
:param PDFObject page: stores the page content of a particular page number
:param int index: stores the page number, in Adobe PDF, page number starts with 0, so 0 is Page 1 
:param Dictionary List processesLinks: Dictionary list containing all the links discovered in the pdf
:param Dictionary List fileMap: Dictionary list containing the external pdf file paths for their respective TFL names
:param Dictionary List oprMap: Dictionary list operation code to be performed on the pdf
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return FilePDFWriter output, PageObject page: returns the modified file output stream and the modified page
"""

def fixExtLinks(output, page, index, processesLinks, fileMap, oprMap, debug=False):

    for key in processesLinks:
        if oprMap['ExternalLink']['OP003'] and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/Launch' and processesLinks[key]['Path'] == None: # rule 1, for removing and fixing external links for TFL names
            try:
                fileMap[processesLinks[key]['Name']] = fileMap[processesLinks[key]['Name']].replace('_', '-')
                fileMap[processesLinks[key]['Name']] = fileMap[processesLinks[key]['Name']].replace('.sas', '-sas.pdf')
                namepath = '../../../../eSub/misc/'+fileMap[processesLinks[key]['Name']]# this can throw error if the key is incorrect, so I am exploiting it to detect invalid keys
                page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)
                output, page = addGOTOR(output, page, namepath, processesLinks[key]['NewerRectCoordinates'], [0,0,0], 1)# 1 is for inherent zoom
            except Exception as e:
                print ("ERROR: *** In the Excel spec sheet, unable to find entry for "+ str(key) + " ***")
                print(traceback.format_exc())
                progExit()
        elif oprMap['TextRemoval']['OP010'] and re.search(r'ad[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
            page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)# deleting the link
            continue
        elif oprMap['TextRemoval']['OP011'] and re.search(r'../[a-z-A-Z0-9]+/ad[a-z-A-Z0-9]+-sas.txt$',processesLinks[key]['Name'], re.M|re.I) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
            page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)# deleting the link
            continue
        elif oprMap['ExternalLink']['OP005'] and (re.search(r't-[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I) or re.search(r'f-[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I) or re.search(r'l-[a-z-A-Z0-9]+.sas$',processesLinks[key]['Name'], re.M|re.I)) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/URI': # rule 2, for removing and fixing external links which have ad*.sas links
            try:
                processesLinks[key]['Path'] = processesLinks[key]['Path'].replace('.txt', '.pdf')
                page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)
                output, page = addGOTOR(output, page, processesLinks[key]['Path'], processesLinks[key]['NewerRectCoordinates'], [0,0,0], 1)
            except Exception as e:
                print ("ERROR: *** while providing file path ***")
                print(traceback.format_exc())
                progExit()
        elif oprMap['ExternalLink']['OP004'] and (re.search(r'ADRG Section [0-9]',processesLinks[key]['Name'], re.M|re.I) or re.search(r'SAP Section [0-9]',processesLinks[key]['Name'], re.M|re.I)) and processesLinks[key]['PageNo'] == index and processesLinks[key]['Type']== '/GoToR': # rule 2, for removing and fixing external links which have ad*.sas links
            try:
                page = removeExtLinks(page, processesLinks[key]['ObjectId'], args.verbose)
                if isinstance(processesLinks[key]['Dest'], list): #bug fixing
                    output, page = addGOTOR(output, page, processesLinks[key]['Path'], processesLinks[key]['NewerRectCoordinates'], [0,0,0], processesLinks[key]['Dest'][0])
                else:
                    output, page = addGOTOR(output, page, processesLinks[key]['Path'], processesLinks[key]['NewerRectCoordinates'], [0,0,0], processesLinks[key]['Dest'])
            except Exception as e:
                print ("ERROR: *** while providing file path ***")
                print(traceback.format_exc())
                progExit()
            
    return output, page

"""
************************************* Method: fixIntLinks *****************************************
Requirements: 
    3.6.    Analysis Results TOC Internal Links
Description:
    Method to fix the Internal Links deviation in the pdf
    It is a little complicated process
    First, go through the processesLinks to fetch the object ID for each internal links which are broken for current index (page number)
    Second, delete those existing links
    Third, match subheadings (subTitle in processesLinks and sublinks in subLinks), if they match then both are same internal links
    Fourth, after successful match, search for the sub-heading's position using findPageByText method
        Fifth, on successful search create the new internal links
    Sixth, On unsuccessful match, point the sublink to the last known page number (partial link, because an internal link with long sentence has been word wrapped in the page)
:param PDFFileWriter instance output: PDF output stream
:param PDFObject page: stores the page content of a particular page number
:param int index: stores the page number, in Adobe PDF, page number starts with 0, so 0 is Page 1 
:param Dictionary List processesLinks: Dictionary list containing all the links discovered in the pdf
:param Dictionary List fileMap: Dictionary list containing the external pdf file paths for their respective TFL names
:param Dictionary List oprMap: Dictionary list operation code to be performed on the pdf
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return FilePDFWriter output, PageObject page: returns the modified file output stream and the modified page
"""

def fixIntLinks(input, output, page, index, processesLinks, subLinks, debug=False):

    pageNo = -1
    titleName = None
    subTitle = None
    pos = 0
    count = 0
    _pages = output.getObject(output._pages)
    for key in processesLinks: # key is the objectID
        if processesLinks[key]['Type'] == '/GoTo' and processesLinks[key]['Path'] != None and processesLinks[key]['PageNo'] == index:
            pageNo = input._getPageNumberByIndirect(processesLinks[key]['Path'][0])
            titleName = processesLinks[key]['Name']
            pos = 0# no need to reset actually
            count= 0
        elif processesLinks[key]['Type'] == '/GoTo' and processesLinks[key]['Path'] == None and processesLinks[key]['PageNo'] == index: #  and str(processesLinks[key]['Name']): # what did I add the last and operator and why am I checking str(processedLinks), need to verify.
            
            try:
                page = removeIntLinks(page, key)
                if debug:
                    print ("NOTE: *** Current sublink from file is "+subLinks[titleName][count]+" ***")
                    print ("NOTE: *** removed broken link from pdf is "+processesLinks[key]['Name']+ " ***")
                subTitle = processesLinks[key]['Name'] # get it from the tsv file
                sublinks = subLinks[titleName][count].split()[0] # get it from the excel sheet and take the first word
                sublinkToMatch = subTitle.split()[0] # take the first word
                if(sublinks.lower() == sublinkToMatch.lower()):#just matching the first words of subtitle and sublink names
                    newPageNo, newPos = findPageByText(titleName, subTitle, input, pos, pageNo, args.verbose)
                    if count < (len(subLinks[titleName])-1):
                        count =  count + 1;
                    if (newPageNo == -1):
                        print ("WARNING: *** Unable to find exact page number for "+subTitle+" ***")
                        pageId = _pages["/Kids"][pageNo].idnum
                        # output.addLink(index, pageNo, processesLinks[key]['NewerRectCoordinates'], border=None)
                        output, page = addGOTO(output, page, pageId, processesLinks[key]['NewerRectCoordinates'], border=None)# unable to find to as a fail safe point to the page number at which figure or table points to
                    else:
                        newpageId = _pages["/Kids"][newPageNo].idnum
                        # output.addLink(index, newPageNo, processesLinks[key]['NewerRectCoordinates'], border=None)
                        output, page = addGOTO(output, page, newpageId, processesLinks[key]['NewerRectCoordinates'], border=None)
                        pageNo = newPageNo # update the last page number, saves some processing
                        pos = newPos
                else:
                    pageId = _pages["/Kids"][pageNo].idnum
                    # output.addLink(index, pageNo, processesLinks[key]['NewerRectCoordinates'], border=None)
                    output, page = addGOTO(output, page, pageId, processesLinks[key]['NewerRectCoordinates'], border=None)
                    if debug:
                        print("NOTE: *** partial link found, linking it to page  "+ str(pageNo)+ " ***")
            except Exception as e:
                print ("ERROR: Exiting, "+str(e))
                print ("ERROR: *** In the Excel spec sheet, unable to find entry for "+ str(titleName)+" ***")
                print(traceback.format_exc())
                progExit()
        
    return output, page

"""
************************************* Method: get_toc *********************************************
Requirements:
    3.2.    Bookmark Parent Child  
    3.4.    ADRG and SAP 
Description:
    Helper method to bookmarkRepair method
    There is a bug in the original getOutlines() method has a bug, because of which bookmark titles are not correct
    This helper method will use another approach to fetch the correct bookmark titles
:param string pdf_path: path to input pdf
:return toc: bookmark titles found in the pdf
    """

def get_toc(pdf_path):

    infile = open(pdf_path, 'rb')
    parser = PDFParser(infile)
    document = PDFDocument(parser)

    toc = list()
    for (level,title,dest,a,structelem) in document.get_outlines():
        toc.append((level, title))

    return toc

"""
************************************* Method: parseContent ****************************************
Requirements: 
    General
Description:
    Method for parsing content Stream to get the blue-colored Links names
    This is used for mapping links found in the /Annots with their respective linknames
:param PageObject page: stores the content of the page provided
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return string list linkNames: contains names of all the link names found in the content stream sequentially
"""

def parseContent(page, debug=False):

    content = page['/Contents'].getObject()
    content = ContentStream(content, page)#creating Contentstream class instance
    linkNames = []
    flag = False
    for operands,operator in content.operations:
        try:
            if operator == b_('rg') and operands == [0,0,1]:#links only
                flag = True
                text = operands[0]
            if operator == b_('Tj') and flag: # first parameter, word to be searched, second parameter, word series were word may be present
                # text = operands[0]
                flag = False
                linkNames.append(operands[0])

        except:
            print ('WARNING: *** problem in deciphering the stream page ***')
    #add some error control here, what if nothing is found
    return linkNames

"""
************************************* Method: progExit ********************************************
Requirements: 
    General
Description:
    Method to safely exit the program in case there is some fatal error
:returns: void
"""
def progExit():
    sys.stdout.flush()
    try:
        os.remove(extLinks) # i will comment this portion in order to not delete this extremely valuable debugging file
    except:
        pass
    sys.exit()

"""
************************************* Method: readConfExcel ***************************************
Requirements: 
    General
Description:
    Method for parsing the Config file
    It is used for debugging purpose
    For testing different operations, suc as text removal, bookmark repair etc
    Please visit the config.xlsx file to get more clear picture
:param string filename: path to Config file
:return Dictionary list oprMap: fetches available operations codes with respective true/false values
"""

def readConfExcel(filename = None):

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active
    oprMap = {}# dictionary for storing the operations
    rows = ws.rows
    for row in rows:

        if (row[1].value!=None and row[2].value!='Category'):
            if (str(row[2].value).strip() not in oprMap): #row[2] is reading value from third column
                oprMap[str(row[2].value).strip()] = {}
            if(str(row[4].value).strip() == 'Yes'):
                oprMap[str(row[2].value).strip()][str(row[1].value).strip()] = True # True means the operation will be executed on pdf
            else:
                oprMap[str(row[2].value).strip()][str(row[1].value).strip()] = False
    return oprMap

"""
************************************* Method: readSpecExcel ***************************************
Requirements: 
    3.5.    Relative External Links
    3.8.    Table External Links                 
    3.10.   Update TFL Program External Links     
    3.11.   Body ADRG and SAP Links 
Description:
    Method for parsing the ANASPEC file
    It fetches TFL names and their respective pdf files
    It also fetched TFL names and their respective sub-headings
:param string filename: path to ANASPEC file
:return Dictionary lists filemap, subLinks: stores the TFL names as key and respective file paths and sub-headings as values
"""

def readSpecExcel(filename = None):

    filemap = {}# dictionary for storing the file-mapping, TFL with file paths
    subLinks = {}# dictionary for storing the sublinks/sub-headings for the Table and Figures
    try:
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb.active

        rows = ws.rows
        for row in rows:

            if (row[0].value!=None):
                if (str(row[0].value).strip() not in subLinks):
                    subLinks[str(row[0].value).strip()] = []
                filemap[str(row[0].value).strip()] = str(row[3].value).strip()
                subLinks[str(row[0].value).strip()].append(str(row[6].value).strip())
    except Exception as e:
        print ("ERROR: *** ANASPEC file is missing in the working directory provided by the User ***")
        progExit()
    return filemap, subLinks

"""
************************************* Method: removeExtLinks **************************************
Requirements: 
    3.5.    Relative External Links
    3.8.    Table External Links                 
    3.9.    Remove ADaM Programs (ad*.sas) Links  
    3.10.   Update TFL Program External Links     
    3.11.   Body ADRG and SAP Links 
Description:
    Method for removing external links in the given page
    If object id of the external links is provided then that only will be deleted, else all external links will be deleted
:param PageObject page: stores the content of the page provided
:param int objectId: stores the object id of the link to be deleted
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return PageObject page: returns the modified page
"""

def removeExtLinks(page, objectId=None, debug=False):

    if "/Annots" in page:
        temp = page["/Annots"][:]
        if (objectId == None):#delete all the external links
            del page['/Annots']
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                obj = annot.getObject()
                if(obj['/A']['/S'] == '/GoToR' or obj['/A']['/S'] == '/Launch' or obj['/A']['/S'] == '/URI'):#look for external links only
                    if debug:
                        print ("NOTE: *** Deleted object id "+str(annot.idnum)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
            if debug:
                print ("NOTE: *** Deleted all the external links ***")
        else:#delete the particular indirect object, or the external link
            del page["/Annots"]
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                if (annot.idnum == objectId):
                    if debug:
                        print ("NOTE: *** Deleted object id "+str(objectId)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
    return page

"""
************************************* Method: removeIntLinks **************************************
Requirements: 
    3.6.    Analysis Results TOC Internal Links
Description:
    Method for removing internal links in the given page
    If object id of the external links is provided then that only will be deleted, else all external links will be deleted
:param PageObject page: stores the content of the page provided
:param int objectId: stores the object id of the link to be deleted
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:return PageObject page: returns the modified page
"""

def removeIntLinks(page, objectId=None, debug=False): 

    if "/Annots" in page:
        temp = page["/Annots"][:]
        if (objectId == None):#delete all the external links
            del page['/Annots']
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                obj = annot.getObject()
                if(obj['/A']['/S'] == '/GoTo'):#look for internal links only
                    if debug:
                        print ("NOTE: *** Deleted object id "+str(annot.idnum)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
            if debug:
                print ("NOTE: *** Deleted all the external links ***")
        else:#delete the particular indirect object, or the external link
            del page["/Annots"]
            page[NameObject('/Annots')] = ArrayObject()
            for annot in temp:
                if (annot.idnum == objectId):
                    if debug:
                        print ("NOTE: *** Deleted object id "+ str(objectId)+" ***")
                    continue
                else:
                    page[NameObject('/Annots')].append(annot)
                    pass
    return page

"""
************************************* Method: removeText ******************************************
Requirements: 
    3.1.    Insert Page X of Y  
    3.3.    Remove Prefixes   
    3.9.    Remove ADaM Programs (ad*.sas) Links 
    3.14.   Program Code Fonts
    3.15.   Remove “AbbVie” from Title Page  
Description:
    Looks for particular pattern in the content of the page and then applies the alternate text
:param PageObject page: object representation of pdf's page
:param string list patterns: contains user defined list of regex patterns
:param string list altTexts: contains user defined list of strings which replace 
    the identified text string via regex pattern
:param boolean debug: initially set to false, set it to true if debug messaged are to be printed
:param boolean ignoreByteStringObject: optional parameter
    to ignore ByteString Objects.
:return PageObject pageRef: returns the modified page
"""

def removeText(page, patterns, altTexts, debug=False, ignoreByteStringObject=False):
    pageRef = page
    content = pageRef['/Contents'].getObject()
    Content = ContentStream(content, pageRef)#creating Contentstream class instance
    for count in range(0, len(Content.operations)):
        operands, operator = Content.operations[count]
        try:
            for index in range(0, len(patterns)):
                pattern = patterns[index]
                altText = altTexts[index]
                if operator == b_('Tj') and re.search(pattern, operands[0], re.M|re.I):
                    text = operands[0]
                    if altText == None:#just strip the link, do not delete the link text
                        replaceString =  text
                        temp, others  = Content.operations[count-1]
                        temp[2] = NumberObject(0)# changing rg array [0,0,1] to [0,0,0]
                    elif altText == '':
                        replaceString =  altText
                        temp, others  = Content.operations[count-3] # 3 obtained by checking the pdf file, can be improved, 
                        temp[0] = TextStringObject(temp[0].replace('(', ''))# deleting the parentheses (, if it is there then it will replace, else wont touch.
                        temp, others  = Content.operations[count+3]
                        temp[0] = TextStringObject(temp[0].replace(')', ''))# deleting the parentheses (
                    elif pattern == r'\bPage \d+$':
                        replaceString = text+altText
                    elif pattern == r'\bProgramming Statements$' and count+10 < len(Content.operations):
                        temp, others  = Content.operations[count+10]
                        if temp[0] == '/F1':
                            temp[1] = NumberObject(9)# changing font size to 9
                        continue

                    else:
                        replaceString = re.sub(pattern, altText, text)# looking for *-Table occurrence in the string #replace a patters with given string
                    if debug:
                        print ("NOTE: *** Replaced string is "+replaceString+" for pattern "+pattern+" ***")
                    if not ignoreByteStringObject:
                        if isinstance(text, TextStringObject):
                            operands[0] = TextStringObject(replaceString)
                    else:
                        if isinstance(text, TextStringObject) or isinstance(text, ByteStringObject):
                            operands[0] = TextStringObject(replaceString)
        except Exception as e:
            print ('WARNING: problem in deciphering the stream page' +str(e))
            print(traceback.format_exc())

    pageRef.__setitem__(NameObject('/Contents'), Content)
    return pageRef


"""
****************************************************************************************************
Operation code for deviation fixing
Set any of these to false to disable that particular operations, settings are provided further below in the code
OpId    OpCode      Category        Operation
1       OP001       Bookmark        Bookmark repair
2       OP002       Bookmark        Add Supplement bookmarks; ADRG and SAP, dependent on Bookmark repair operation
3       OP009       TextRemoval     Fix page numbers, add Page X of Total-Pages
4       OP010       TextRemoval     Remove ad*.sas link color under Analysis Datasets
5       OP011       TextRemoval     Delete ad*-sas.txt links
6       OP013       TextRemoval     Remove AbbVie name
7       OP003       ExternalLink    External Links to TFL Summary Data PDFs
8       OP004       ExternalLink    Adding destination page numbers for External PDFs in SAP and ADRG
9       OP005       ExternalLink    External Links to TFL*-sas.pdf in place of txt files
10      OP006       InternalLink    Fix internal links under Analysis Result Metadata
11      OP007       TextRemoval     Fix Table names, remove prefixes
12      OP008       TextRemoval     Fix Figure names, remove prefixes
13      OP012       TextRemoval     Change the font of Code snippet to Normal font

************************************** Main Execution begin ****************************************
"""
# parsing the arguments passed                              
# setting the log file destination
print ("\n*************************************** pdftoolkit started ***************************************\n")
if args.logFile!="":
    sys.stdout = open(args.logFile, "w")

if(not (re.search(r'.pdf', args.inFile, re.M|re.I) and re.search(r'.pdf', args.outFile, re.M|re.I))):
    print ("ERROR: *** Cannot find .pdf extension ***")
    print ("Please mention the input and output pdf file names in proper format, -h for more info")
    print ("first argument is input file and second argument is output file")
    print ("For example.. ")
    print ("python main.py -i path/input.pdf -o path/output.pdf")
    progExit()

#Checking if the ANASPEC file has been passed by the user
if args.specFile: # if passed then check below for extension
    if(not (re.search(r'.xlsx', args.specFile, re.M|re.I))): #check if the file has proper extension format.
        print ("ERROR: *** Cannot find .xlsx extension for the spec file ***")
        print ("-h or --help for more running the script in proper format")
        print ("Exiting now")
        progExit()

input = PdfFileReader(open(args.inFile, "rb")) # reading the input pdf file
print ("NOTE: *** You are currently running Python PDFTools version 1.0 ***")
sys.stdout.flush()
NumofPages = input.getNumPages()
print ("NOTE: *** %d pages processed Successfully ***"%(NumofPages))
sys.stdout.flush()
print ("NOTE: *** Processing Document "+ args.inFile +" ***")
sys.stdout.flush()
outlines = input.getOutlines() # fetch bookmarks
titles = get_toc(args.inFile) # fetch bookmarks using alternate method, using pdfminer, this gives bug free bookmark names

"""
****************************************************************************************************
************************************** Creating rules **********************************************
"""
#checking if the CONFIG file has been passed by the user
if args.configFile: # if passed then check below for extension
    if(not (re.search(r'.xlsx', args.configFile, re.M|re.I))): #check if the file has proper extension format.
        print ("ERROR: *** Cannot find .xlsx extension for the config file ***")
        print ("-h or --help for more running the script in proper format")
        print ("Exiting now")
        progExit()
    print ("NOTE: *** Fetching operation list from the Configuration spreadsheet ***")
    try:
        oprMap = readConfExcel(filename =args.configFile)
    except Exception as e:
        print ('ERROR: *** Some error occurred while reading the config spreadsheet ***')
        print(traceback.format_exc())
else:#if CONFIG file not provided by the user then create the rules/operations yourselves on the basis of ANASPEC file
    #default operations to be performed even if the ANASPEC file is present or not, True means operation to be followed
    oprMap = {'Bookmark': {'OP002': True, 'OP001': True}, 'TextRemoval': {'OP008': False, 'OP009': True, 'OP011': True, 'OP010': True, 'OP012': False, 'OP007': False, 'OP013': True}, 'InternalLink': {'OP006': False}, 'ExternalLink': {'OP003': False, 'OP004': False, 'OP005': False}}
    if checkForAnaspec(outlines):
        if args.specFile: #if ANASPEC file is there then add following rules/operations
            #setting following rules for ANASPEC
            oprMap['TextRemoval']['OP007'] =  True
            oprMap['TextRemoval']['OP008'] =  True
            oprMap['TextRemoval']['OP012'] =  True
            oprMap['ExternalLink']['OP003'] =  True
            oprMap['ExternalLink']['OP004'] =  True
            oprMap['ExternalLink']['OP005'] =  True
            oprMap['InternalLink']['OP006'] =  True
        else:
            print ("")
            print ("")
            print ("ERROR: *** The PDF document you provided needs ANASPEC file, please provide this option and try again ***")
            sys.stdout.flush()
            progExit()
    else:
        if args.specFile: #if ANASPEC file is there but is actually not needed
            print ("")
            print ("")
            print ("ERROR: *** The input pdf file does not require an Analysis Result Specification Excel.  Remove this option and try again. ***")
            sys.stdout.flush()
            progExit()


# creating TextRemoval rule
patterns = []
altTexts = []
if(oprMap['TextRemoval']['OP007']):
    patterns.append( r'^\d*[0-9]-Table ')
    altTexts.append('Table ')
if(oprMap['TextRemoval']['OP008']):
    patterns.append(r'^\d*[0-9]-Figure ')
    altTexts.append('Figure ')
if(oprMap['TextRemoval']['OP009']):
    patterns.append( r'\bPage \d+$')
    altTexts.append(' of '+str(NumofPages))
if(oprMap['TextRemoval']['OP010']):
    patterns.append(r'\bad[a-z-A-Z0-9]+.sas$')
    altTexts.append(None)
if(oprMap['TextRemoval']['OP011']):
    patterns.append(r'../[a-z-A-Z0-9]+/ad[a-z-A-Z0-9]+-sas.txt$')
    altTexts.append('')
if(oprMap['TextRemoval']['OP012']):
    patterns.append(r'\bProgramming Statements$')
    altTexts.append('Programming Statements')
if(oprMap['TextRemoval']['OP013']):
    patterns.append(r'\bAbbvie$')
    altTexts.append(' ')

"""
*****************************************************************************************************
"""

processesLinks = [] # for getting all the links present in the pdf
fileMap = [] # for getting dictionary list of TFL names along with their pdf file paths
subLinks = [] # for getting dictionary list of TFL names with their sub-headings

"""
*****************************************************************************************************
************************************** reading from ANASPEC file ************************************
"""
if (oprMap['InternalLink']['OP006'] or oprMap['ExternalLink']['OP003']):# only this two opcodes actually needs ANASPEC file, hence if these are active then only read the file
    print ("NOTE: *** Getting TFL mapping from the ANASPEC Excel file ***")
    sys.stdout.flush()
    print ("NOTE: *** TFL mapped Successfully ***")
    sys.stdout.flush()
    fileMap, subLinks = readSpecExcel(filename =args.specFile)# fetch correct file-mapping from the given excel file, ANASPEC sheet
    sys.stdout.flush()
"""
*****************************************************************************************************
************************************** Fetching all the links from pdf ******************************
"""
if (oprMap['TextRemoval']['OP010'] or oprMap['TextRemoval']['OP011'] or oprMap['InternalLink']['OP006'] or oprMap['ExternalLink']['OP005'] or oprMap['ExternalLink']['OP004'] or oprMap['ExternalLink']['OP003']):
    print ("NOTE: *** Identifying the all the external and internal links in the PDF doc ***")
    sys.stdout.flush()
    processesLinks = findLink(input, args.verbose)# find all external links from the pdf and store them in a dictionary list for further processing
    print ("NOTE: *** Links identified Successfully ***")
    sys.stdout.flush()

"""
*****************************************************************************************************
********** Processing pages and correcting errors like incomplete links, wrong words etc*************
"""
output = PdfFileWriter() ##Opening the output stream
print ("NOTE: *** Performing corrections on the PDF ***")
sys.stdout.flush()
for index in range(0,NumofPages):
        if args.logFile=="" and False:
            sys.stdout.write("\rNOTE: *** Successfully processed: %d of %d pages " %(index+1,NumofPages))
            sys.stdout.flush()
        page = input.getPage(index)
        if(len(patterns)!=0):
            page = removeText(page, patterns, altTexts, args.verbose)# fixing the incorrect Figure names, table names and adding Page numbers
        output, page = fixExtLinks(output, page, index, processesLinks, fileMap, oprMap,  args.verbose)
        page.compressContentStreams() # compress the content stream using the Flat-encode, decreases the size of the document by 3/4
        output.addPage(page)
if oprMap['InternalLink']['OP006']:
    print ("\nNOTE: *** Correcting Internal Links of the PDF ***")
    for index in range(0,NumofPages):
        page = output.getPage(index)            
        output, page = fixIntLinks(input, output, page, index, processesLinks, subLinks, args.verbose) # can be moved into upper for loop
"""
*****************************************************************************************************
************************************** All correction done ******************************************
"""
if args.logFile!="" and False:
    print ("\nNOTE: *** Corrections applied successfully on all %d pages***"%(NumofPages))
else:
    print ("\nNOTE: *** Corrections applied successfully on all %d pages***"%(NumofPages))
sys.stdout.flush()
print ("NOTE: *** Repairing bookmarks ***")
sys.stdout.flush()
output = bookMarksRepair(output, outlines, titles, oprMap, args.verbose)
print ("NOTE: *** Bookmarks repaired successfully ***")
sys.stdout.flush()
output.setPageMode("/UseOutlines") #This is what tells the PDF to open with bookmarks at left side
# sys.stdout = sys.__stdout__
try:
    os.remove(extLinks)
except:
    pass

print("NOTE: *** Saving the corrected pdf ***")
while True:
    try:
        outputStream = file(args.outFile, "wb")
        break
    except:
        print ("ERROR: *** Corrected pdf cannot be saved as the "+args.outFile+" file is open, please close it and press Enter to continue ***")
        sys.stdout.flush()
        paused =  raw_input("Thanks \n")

output.write(outputStream)
outputStream.close()
print("NOTE: *** Corrected pdf saved successfully ***")

"""
*****************************************************************************************************
************************************** End of line **************************************************
"""